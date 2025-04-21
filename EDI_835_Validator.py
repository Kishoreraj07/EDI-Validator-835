import sys
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                            QPushButton, QTextEdit, QFileDialog, QLineEdit, QMessageBox,
                            QLabel, QFrame, QSplitter, QStatusBar, QAction,
                            QSizePolicy)
from PyQt5.QtGui import QTextCharFormat, QColor, QIcon, QFont, QPalette, QPixmap, QTextCursor, QKeySequence
from PyQt5.QtCore import Qt, QFileSystemWatcher, QSize, QCoreApplication
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# def validation_singlefile(filepath):
#     return [[{'lines': 5, 'value': '28310_MSMB_HB_20241212100524_DEVOTED_1348543_12-09-24_$44453.76', 'position': 2, 'field': 'Header', 'title': 'payer_ref_no'}], 
#             [{'lines': 371, 'value': '', 'position': 5, 'field': 'CLP', 'title': 'pat_middlename'}], 
#             [{'lines': 372, 'value': '', 'position': 5, 'field': 'CLP', 'title': 'pat_middlename'}], 
#             [{'lines': 16, 'value': '-289.41', 'position': 3, 'field': 'ServiceLine', 'title': 'adjustment_amt'}]]

# def validation_multiplefile(folder_path, excel_path):
#     return True


def validation_singlefile(filepath):
    def split_tilde(data):
        if '~' in data:
            data=data.split('~')[0]
            return data
        elif '~' not in data:
            return data
    def validation(title,data):
        if title in ['sender_id','rec_id','payer_id']:
            if re.fullmatch(r'^[a-zA-Z0-9]{1,15}$', data): #aplha numeric max 1 - 15 
                return True
            return False
        elif title in ['auth_info_qualifier']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9]{2}$", data))
        elif title in ['authorization_info']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9 ]{10}$", data))
        elif title in ['interchange_sender_id']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9 ]{15}$", data))
        elif title in ['interchange_date']:
            return bool(re.fullmatch(r"^\d{2}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$", data))
        elif title in ['interchange_time']:
            return bool(re.fullmatch(r"^([01][0-9]|2[0-3])[0-5][0-9]$", data))
        elif title in ['interchange_standard_id']:
            return bool(re.fullmatch(r"^.{1}$", data))
        elif title in ['transcation_heading_code']:
            if data in 'CDHIPUX':
                return True
            else:
                return False
        elif title in ['pay_format_code']:
            if data in ['CCP','']:
                return True
            else:
                return False
        elif title in ['ref_identification_qul']:
            if data =='EV':
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,2}$", data))
        elif title in ['id_no_qualifier']:
            if data in ['01','']:
                return True
            else:
                return False
            
        elif title in ['identification_bpr']:
            if data in ['999999999','']:
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9]{3,12}$", data))
        
        elif title in ['account_number']:
            if data in ['']:
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,35}$", data))            
        elif title in ['effective_entry_date']:
            pattern = r"^(19|20)\d{2}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['included_sgmt_no']:
            return bool(re.fullmatch(r"^\d{1,10}$", data))
        elif title in ['number_of_transaction_set']:
            return bool(re.fullmatch(r"\d{1,6}", data))
        elif title in ['group_control_no']:
            return bool(re.fullmatch(r"\d{1,9}", data))
        elif title in ['no_of_included_segments']:
            return bool(re.fullmatch(r"\d{1,5}", data))
        elif title in ['interchange_control_no']:
            return bool(re.fullmatch(r"\d{9}", data))
        elif title in ['trans_set_control_number']:
            return bool(re.fullmatch(r"^\d{4,9}$", data))
        elif title in ['acc_no_qualifier_code']:
            if data in ['DA','']:
                return True
            else:
                return False
        elif title in ['originating_company_identfier']:
            if data in ['9999999999','']:
                return True
            else:
                pattern = r"^[a-zA-Z0-9]{10}$"
                return bool(re.fullmatch(pattern, data))
                
        elif title in ['originating_supplemental_code']:
            if data in ['','199999999']:
                return True
            else:
                pattern = r"^[a-zA-Z0-9]{9}$"
                return bool(re.fullmatch(pattern, data))
            
        elif title in ['rdfi_id_number']:
            if data in ['','01']:
                return True
            else:
                return False
        elif title in ['rdfi_identification_number']:
            if data in ['','999999999']:
                return True
            else:
                pattern = r"^[a-zA-Z0-9]{3,12}$"
                return bool(re.fullmatch(pattern, data))
        
        elif title in ['cre_deb_flag_code']:
            if data in ['C','D']:
                return True
            else:
                return False
        elif title in ['intchangecontrolno']:
            return bool(re.fullmatch(r"\d{9}", data)) #numeric 9 digit
        elif title in ['grp_con_no','tans_set_con_no']:
            return bool(re.fullmatch(r'^\d{1,9}$', data)) #numeric max 9 digit
        elif title in ['total_pay_amt']:
            pattern = r'^-?\d{1,18}(\.\d{1,2})?$'
            return bool(re.fullmatch(pattern, data))
        elif title in ['pay_method_code']:
            if data in ['ACH', 'CHK', 'NON']:
                return True
            else:
                return False
        elif title in ['bank_acc_no']:
            return bool(re.fullmatch(r'^[a-zA-Z0-9]{1,17}$', data))
        elif title in ['check_eft_trace_no']:
            return bool(re.fullmatch(r'^[a-zA-Z0-9]{1,20}$', data))
        elif title in ['payer_city','payee_city']:
            return bool(re.fullmatch(r'^[a-zA-Z0-9 ]{1,30}$', data))
        elif title in ['trace_type_code']:
            if data =='1':
                return True
            else:
                return False
        elif title in ['check_no']:
            return bool(re.fullmatch(r'^[a-zA-Z0-9]{1,30}$', data))
        elif title in ['payer_id','payee_id']:
            # Validates if the input is alphanumeric and at most 15 characters long.
            return bool(re.fullmatch(r'^[a-zA-Z0-9]{1,15}$', data))
        elif title in ['curr_code']:
            if data in ['USD','CAD']:
                return True
            else:
                return False
        elif title in ['pay_ref_qul']:
            if len(data)==2 and data in ['EV']:
                return True
            else:
                return False
        elif title in ['ref_identification_num']:
            return bool(re.fullmatch(r"^.{1,30}$", data))
        elif title in ['date_aualifier']:
            if data in ['405']:
                return True
            else:
                return False
        elif title in ['service_date_qualifier']:
            if title in ['472','150','151']:
                return True
            else:
                return bool(re.fullmatch(r"^\d{3}$", data))
        elif title in ['pay_date']:
            # Validates if the input is an 8-digit date in YYYYMMDD format.
            pattern = r"^(19|20)\d{2}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['payer_name','payee_name']:
            # Validates if the input is alphanumeric and at most 60 characters long.
            return bool(re.fullmatch(r'^[a-zA-Z0-9 .\-]{1,60}$', data))
        elif title in ['payee_id_code_qualifier']:
            if data in ['FI','XX']:
                return True
            else:
                return False
        elif title in ['payer_address','payee_address']:
            # Validates if the input is alphanumeric and at most 55 characters long.
            return bool(re.fullmatch(r'^[a-zA-Z0-9 \-]{1,55}$', data))
        elif title in ['payer_state','payee_state']:
            # Validates if the input is exactly 2 uppercase letters (US State Code).
            return bool(re.fullmatch(r'^[A-Z]{2}$', data))
        elif title in ['payer_zip','payee_zip']:
            # Validates if the input is a 5-digit or 9-digit ZIP code (ZIP+4 format).
            return bool(re.fullmatch(r'^\d{5}(\d{4})?$', data))
        elif title in ['service_line_no']:
            # Validates if the input is a numeric value with 1 to 9 digits.
            return bool(re.fullmatch(r'^\d{1,9}$', data))
        elif title in ['claim_no']:
            # Validates if the input is alphanumeric with a maximum of 38 characters.
            return bool(re.fullmatch(r'^[a-zA-Z0-9]{1,38}$', data))
        elif title in ['claim_sts_code']:
            # Validates if the input is a numeric value with 1 to 4 digits.
            return bool(re.fullmatch(r'^\d{1,2}$', data))
        elif title in ['total_charges','total_pay_amt']:
            # Decimal, max 18 digits, 2 decimal places
            return bool(re.fullmatch(r'^-?\d{1,18}(\.\d{1,2})?$', data))
        elif title in ['clm_indicator_code']:
            pattern = r"^[a-zA-Z0-9]{1,2}$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['payer_claim_control_no']:
            pattern = r"^[a-zA-Z0-9]{1,30}$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['pat_lastname']:
            return bool(re.fullmatch(r'^[a-zA-Z ]{1,35}$', data))
        elif title in ['entity_type_qualifier']:
            if data == '1':
                return True
            else:
                return False
        elif title in ['pat_firstname']:
            return bool(re.fullmatch(r'^[a-zA-Z ]{1,25}$', data))
        elif title in ['pat_middlename']:
            if data=='':
                return True
            else:
                pattern = r"^[a-zA-Z0-9]{1,25}$"
                return bool(re.fullmatch(pattern, data))
        elif title in ['pat_suffix','pat_prefix']:
            if data == '':
                return True
            pattern = r"^[a-zA-Z0-9]{1,10}$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['provider_lastname']:
            if data == '':
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9 ]{1,60}$", data))
        elif title in ['provider_firstname']:
            if data == '':
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9 ]{1,35}$", data))
        elif title in ['provider_middlename']:
            if data == '':
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,25}$", data))
        elif title in ['provider_prefix']:
            if data == '':
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,10}$", data))
        elif title in ['provider_id_code_qualifier']:
            if data in ['XX','FI']:
                return True
            else:
                return False
        elif title in ['provider_id_code']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9]{2,80}$", data))
        elif title in ['clp_id_code']:
            pattern = r"^[a-zA-Z0-9]{2,80}$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['id_code_qualifier']:
            pattern = r"^[a-zA-Z0-9]{1,2}$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['procedure_code']:
            if data in ['HC','NU']:
                return True
            else:
                return False
        elif title in ['product_service_id']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,48}$", data))
            
        elif title in ['service_charge_amt','service_pay_amount','paid_amount']:
            if data in  ['0','']:
                return True
            else:
                return bool(re.fullmatch(r"^-?\d{1,18}(\.\d{1,2})?$", data))

        elif title in ['claim_grp_code']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,2}$", data))
        elif title in ['adjustment_reason_code']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,5}$", data))
        elif title in ['adjustment_amt']:
            return bool(re.fullmatch(r"^-?\d{1,18}(\.\d{1,2})?$", data))

        


    def chunk_file_content(file_content):
        # Split the file content into lines
        lines = file_content.split('\n')
        
        # Identify lines that start with 'CLP*'
        clp_lines = [i for i, line in enumerate(lines) if line.startswith("CLP*")]
        
        # Identify the header data (lines before the first CLP*)
        header_data = lines[:clp_lines[0]] if clp_lines else []

        # Initialize a list for each CLP data chunk
        clp_data_chunks = []
        
        # Loop through each CLP, and chunk data accordingly
        for i in range(len(clp_lines)):
            start = clp_lines[i]
            if i + 1 < len(clp_lines):
                end = clp_lines[i + 1]
            else:
                # For the last CLP, we want to go up to the first IEA* line
                iea_line = next((j for j, line in enumerate(lines) if line.startswith("IEA*")), len(lines))
                end = iea_line
            
            clp_data_chunks.append((f"CLP_{i + 1}", lines[start:end]))
        clpmaindata=[]
        for indexx,clpd in enumerate(clp_data_chunks):
            clpmaindata.append(clpd[1])
        # clpdata=clp_data_chunks[1]
        svc_final_chunk=[]
        for indexxx,clpdata in enumerate(clpmaindata):
            svc_chunk=[]
            svc_fd=False
            
            for indx4,svcchunk in enumerate(clpdata):
                if 'SVC*' in svcchunk:
                    svc_fd=True
                if svcchunk.startswith("SE*"):
                    break
                if svc_fd:
                    svc_chunk.append(svcchunk)
            svc_final_chunk.append(svc_chunk)
        clp_data_final=[]
        for clpdata in clpmaindata:
            clp_data=[]
            for clpdata2 in clpdata:
                if clpdata2.startswith('SVC*'):
                    break
                else:
                    clp_data.append(clpdata2)
            clp_data_final.append(clp_data)
        return header_data, clp_data_chunks,svc_final_chunk,lines,clp_data_final
    # for file_name in files:
    failed_info=[]
    file_path=filepath
    # file_path2=os.path.dirname(file_path)
    # new_path=os.path.join(file_path2,'temp.835')
    # shutil.copy(file_path,new_path)
    with open(file_path, "r", encoding="utf-8") as f1:
            file_content = f1.read()
            # f1.close()
    header_data, clp_chunks,svc_chunks,full_lines,clp_full = chunk_file_content(file_content)
    total_lines_count=len(full_lines)
    header_count=len(header_data)
    clp_count=0
    clpcount2=0
    for clpfu in clp_full:
        clpcount2+=len(clpfu)
    svc_count2=0
    for svcfu in svc_chunks:
        svc_count2+=len(svcfu)
    totalcount2=header_count+clpcount2+svc_count2
    for clpd1 in clp_full:
        clp_count+=len(clpd1)
    svc_count=0
    for svcd1 in svc_chunks:
        svc_count+=len(svcd1)
    for index1,header in enumerate(header_data):
        if header.startswith('ISA*'):
            hp1=header.split('*')
            # for ind_,hp1_t in enumerate(hp1):
            #     hp1[ind_]=hp1_t.strip()
            auth_info_qualifier=validation('auth_info_qualifier',split_tilde(hp1[1]))
            authorization_info=validation('authorization_info',split_tilde(hp1[2]))
            security_info_qualifier=validation('auth_info_qualifier',split_tilde(hp1[3]))
            security_information=validation('authorization_info',split_tilde(hp1[4]))
            interchange_sender_id=validation('interchange_sender_id',split_tilde(hp1[6]))
            interchange_id_qualifier=validation('auth_info_qualifier',split_tilde(hp1[7]))
            interchange_receiver_id=validation('interchange_sender_id',split_tilde(hp1[8]))
            interchange_date=validation('interchange_date',split_tilde(hp1[9]))
            interchange_time=validation('interchange_time',split_tilde(hp1[10]))
            interchange_standard_id=validation('interchange_standard_id',split_tilde(hp1[11]))

            # sender_id_validate=validation('sender_id',hp1[6])
            # rec_id_validate=validation('rec_id',hp1[8])
            interchange_controlno_validate=validation('intchangecontrolno',hp1[13])
            validation_results = {
                'auth_info_qualifier': (auth_info_qualifier, split_tilde(hp1[1]), 1),
                'authorization_info': (authorization_info, split_tilde(hp1[2]), 2),
                'security_info_qualifier': (security_info_qualifier, split_tilde(hp1[3]), 3),
                'security_information': (security_information, split_tilde(hp1[4]), 4),
                'interchange_sender_id': (interchange_sender_id, split_tilde(hp1[6]), 6),
                # 'sender_id': (sender_id_validate, hp1[6], 6),
                'interchange_id_qualifier': (interchange_id_qualifier, split_tilde(hp1[7]), 7),
                'interchange_receiver_id': (interchange_receiver_id, split_tilde(hp1[8]), 8),
                'interchange_date': (interchange_date, split_tilde(hp1[9]), 9),
                'interchange_time': (interchange_time, split_tilde(hp1[10]), 10),
                'interchange_standard_id': (interchange_standard_id, split_tilde(hp1[11]), 11),
                'intchangecontrolno': (interchange_controlno_validate, hp1[13], 13),
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('GS*'):
            hp2=header.split('*')
            sender_id2_vaidate=validation('sender_id',hp2[2])
            rec_id2_validate=validation('rec_id',hp2[3])
            grp_control_number_validate=validation('grp_con_no',hp2[6])
            validation_results = {
                'sender_id2': (sender_id2_vaidate, hp2[2],2),
                'rec_id2': (rec_id2_validate, hp2[3],3),
                'grp_control_number': (grp_control_number_validate, hp2[6],6)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('ST*835'):
            hp3=header.split('*')
            trans_set_con_no_validate=validation('tans_set_con_no',split_tilde(hp3[2]))
            validation_results = {
                'trans_set_con_no': (trans_set_con_no_validate, split_tilde(hp3[2]),2)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('BPR*'):
            hp4=header.split('*')
            transcation_heading_code=validation('transcation_heading_code',split_tilde(hp4[1]))
            total_pay_amt=validation('total_pay_amt',hp4[2])
            cre_deb_flag_code=validation('cre_deb_flag_code',split_tilde(hp4[3]))
            pay_method_code=validation('pay_method_code',split_tilde(hp4[4]))
            pay_format_code=validation('pay_format_code',split_tilde(hp4[5]))
            id_no_qualifier=validation('id_no_qualifier',split_tilde(hp4[6]))
            identification_bpr=validation('identification_bpr',split_tilde(hp4[7]))
            acc_no_qualifier_code_bpr1=validation('acc_no_qualifier_code',split_tilde(hp4[8]))
            account_number_bpr1=validation('account_number',split_tilde(hp4[9]))
            originating_company_identfier=validation('originating_company_identfier',split_tilde(hp4[10]))
            originating_supplemental_code=validation('originating_supplemental_code',split_tilde(hp4[11]))
            rdfi_id_number=validation('rdfi_id_number',split_tilde(hp4[12]))
            rdfi_identification_number=validation('rdfi_identification_number',split_tilde(hp4[13]))
            account_no_qual_code_bpr2=validation('acc_no_qualifier_code',split_tilde(hp4[14]))
            account_number_bpr2=validation('account_number',split_tilde(hp4[15]))
            effective_entry_date=validation('effective_entry_date',split_tilde(hp4[16]))
            
            validation_results = {
                'transcation_heading_code': (transcation_heading_code, split_tilde(hp4[1]), 1),
                'total_pay_amt': (total_pay_amt, hp4[2], 2),
                'cre_deb_flag_code': (cre_deb_flag_code, split_tilde(hp4[3]), 3),
                'pay_method_code': (pay_method_code, split_tilde(hp4[4]), 4),
                'pay_format_code': (pay_format_code, split_tilde(hp4[5]), 5),
                'id_no_qualifier': (id_no_qualifier, split_tilde(hp4[6]), 6),
                'identification_bpr': (identification_bpr, split_tilde(hp4[7]), 7),
                'acc_no_qualifier_code_bpr1': (acc_no_qualifier_code_bpr1, split_tilde(hp4[8]), 8),
                'account_number_bpr1': (account_number_bpr1, split_tilde(hp4[9]), 9),
                'originating_company_identfier': (originating_company_identfier, split_tilde(hp4[10]), 10),
                'originating_supplemental_code': (originating_supplemental_code, split_tilde(hp4[11]), 11),
                'rdfi_id_number': (rdfi_id_number, split_tilde(hp4[12]), 12),
                'rdfi_identification_number': (rdfi_identification_number, split_tilde(hp4[13]), 13),
                'account_no_qual_code_bpr2': (account_no_qual_code_bpr2, split_tilde(hp4[14]), 14),
                'account_number_bpr2': (account_number_bpr2, split_tilde(hp4[15]), 15),
                'effective_entry_date': (effective_entry_date, split_tilde(hp4[16]), 16)
            }

            failed_entry = [
                {'lines': index1, 'value': value, 'position': position, 'field': 'Header', 'title': var_name}
                for var_name, (is_valid, value, position) in validation_results.items()
                if not is_valid
            ]

            if failed_entry:
                failed_info.append(failed_entry)



        elif header.startswith('TRN*'):
            hp5=header.split('*')
            trace_type_code_validate=validation('trace_type_code',hp5[1])
            check_no_validate=validation('check_no',hp5[2])
            payer_id_validate=validation('payer_id',split_tilde(hp5[3]))
            validation_results = {
                'trace_type_code': (trace_type_code_validate, hp5[1],1),
                'check_no': (check_no_validate, hp5[2],2),
                'payer_id': (payer_id_validate, split_tilde(hp5[3]),3)

            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('CUR*'):
            hp6=header.split('*')
            curr_code_validate=validation('curr_code',split_tilde(hp6[1]))
            validation_results = {
                'curr_code': (curr_code_validate, split_tilde(hp6[1]),1)

            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('REF*') and header_data[index1+1].startswith('LX*'):
            hp16=header.split('*')
            payee_id=validation('payee_id',split_tilde(hp16[2]))
            validation_results = {
                'payee_id': (payee_id, split_tilde(hp16[2]),2)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('LX*'):
            hp17=header.split('*')
            service_line_no=validation('service_line_no',split_tilde(hp17[1]))
            validation_results = {
                'service_line_no': (service_line_no, split_tilde(hp17[1]),1)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('REF*') and (header_data[index1+2].startswith('N1*PR*') or header_data[index1+1].startswith('N1*PR*')):
            hp7=header.split('*')
            ref_identification_qul_validate=validation('pay_ref_qul',split_tilde(hp7[1]))
            ref_identification_num=validation('ref_identification_num',split_tilde(hp7[2]))
            validation_results = {
                'ref_identification_qul': (ref_identification_qul_validate, split_tilde(hp7[1]),1),
                'ref_identification_num': (ref_identification_num, split_tilde(hp7[2]),2)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('DTM*'):
            hp8=header.split('*')
            date_qaualifier=validation('date_aualifier',split_tilde(hp8[1]))
            pay_date=validation('pay_date',split_tilde(hp8[2]))
            validation_results = {
                'date_qaualifier':(date_qaualifier, split_tilde(hp8[1]),1),
                'pay_date': (pay_date, split_tilde(hp8[2]),2)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('N1*PR*'):
            hp9=header.split('*')
            payer_name=validation('payer_name',split_tilde(hp9[2]))
            validation_results = {
                'payer_name': (payer_name, split_tilde(hp9[2]),2)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('N3*') and (header_data[index1-2].startswith('N1*PR*') or header_data[index1-1].startswith('N1*PR*')):
            hp10=header.split('*')
            payer_address=validation('payer_address',split_tilde(hp10[1]))
            validation_results = {
                'payer_address': (payer_address, split_tilde(hp10[1]),1)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('N4*') and (header_data[index1-2].startswith('N1*PR*') or header_data[index1-1].startswith('N1*PR*')):
            hp11=header.split('*')
            payer_city=validation('payer_city',split_tilde(hp11[1]))
            payer_state=validation('payer_state',split_tilde(hp11[2]))
            payer_zip=validation('payer_zip',split_tilde(hp11[3]))
            validation_results = {
                'payer_city': (payer_city, split_tilde(hp11[1]),1),
                'payer_state': (payer_state, split_tilde(hp11[2]),2),
                'payer_zip': (payer_zip, split_tilde(hp11[3]),3)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('REF*') and (header_data[index1+2].startswith('N1*PE*') or header_data[index1+1].startswith('N1*PE*')):
            hp12=header.split('*')
            payer_id=validation('payer_id',split_tilde(hp12[2]))
            validation_results = {
                'payer_id': (payer_id, split_tilde(hp12[2]),2)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('N1*PE*'):
            hp13=header.split('*')
            payee_name=validation('payee_name',split_tilde(hp13[2]))
            payee_id_code_qualifier=validation('payee_id_code_qualifier',split_tilde(hp13[3]))
            validation_results = {
                'payee_name': (payee_name, split_tilde(hp13[2]),2),
                'payee_id_code_qualifier': (payee_id_code_qualifier, split_tilde(hp13[3]),3)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('N3*') and (header_data[index1-2].startswith('N1*PE*') or header_data[index1-1].startswith('N1*PE*')):
            hp14=header.split('*')
            payee_address=validation('payee_address',split_tilde(hp14[1]))
            validation_results = {
                'payee_address': (payee_address, split_tilde(hp13[2]),2)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        elif header.startswith('N4*') and (header_data[index1-2].startswith('N1*PE') or header_data[index1-1].startswith('N1*PE*')):
            hp15=header.split('*')
            payee_city=validation('payee_city',split_tilde(hp15[1]))
            payee_state=validation('payee_state',split_tilde(hp15[2]))
            payee_zip=validation('payee_zip',split_tilde(hp15[3]))
            validation_results = {
                'payee_city': (payee_city, split_tilde(hp15[1]),1),
                'payee_state': (payee_state, split_tilde(hp15[2]),2),
                'payee_zip': (payee_zip, split_tilde(hp15[3]),3)
            }
            failed_entry=[
                {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                for var_name, (is_valid, value,position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        
    for index2,clpd2 in enumerate(clp_full): 
        if index2==0:
            clpline=len(header_data)
        elif index2==1:
            clpline=len(header_data)+len(clp_full[index2-1])+len(svc_chunks[index2-1])
        else:
            clpline=clpline+len(svc_chunks[index2-1])
        for index3,clpdata in enumerate(clpd2):
            if clpdata.startswith('CLP*'):
                cp1=clpdata.split('*')
                claim_no=validation('claim_no',split_tilde(cp1[1]))
                claim_sts_code=validation('claim_sts_code',split_tilde(cp1[2]))
                total_charges=validation('total_charges',split_tilde(cp1[3]))
                total_pay_amt=validation('total_pay_amt',split_tilde(cp1[4]))
                pat_responsibility_amt=validation('total_pay_amt',split_tilde(cp1[5]))
                clm_indicator_code=validation('clm_indicator_code',split_tilde(cp1[6]))
                payer_claim_control_no=validation('payer_claim_control_no',split_tilde(cp1[7]))

                validation_results = {
                'claim_no': (claim_no, split_tilde(cp1[1]),1),
                'claim_sts_code': (claim_sts_code, split_tilde(cp1[2]),2),
                'total_charges': (total_charges, split_tilde(cp1[3]),3),
                'total_pay_amt': (total_pay_amt, split_tilde(cp1[4]),4),
                'pat_responsibility_amt': (pat_responsibility_amt, split_tilde(cp1[5]),5),
                'clm_indicator_code': (clm_indicator_code, split_tilde(cp1[6]),6),
                'payer_claim_control_no': (payer_claim_control_no, split_tilde(cp1[7]),7)
                }
                failed_entry=[
                    {'lines': clpline, 'value': value,'position':position,'field':'CLP','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
                
            elif clpdata.startswith('NM1*') and '*QC*' in clpdata:
                cp2=clpdata.split('*')
                entity_type_qualifier,pat_firstname,pat_lastname,pat_middlename,pat_prefix,pat_suffix,id_code_qualifier,clp_id_code='notfound','notfound','notfound','notfound','notfound','notfound','notfound','notfound'
                try:
                    entity_type_qualifier=validation('entity_type_qualifier',split_tilde(cp2[2]))
                    pat_lastname=validation('pat_lastname',split_tilde(cp2[3]))
                    pat_firstname=validation('pat_firstname',split_tilde(cp2[4]))
                    pat_middlename=validation('pat_middlename',split_tilde(cp2[5]))
                    pat_prefix=validation('pat_prefix',split_tilde(cp2[6]))
                    pat_suffix=validation('pat_suffix',split_tilde(cp2[7]))
                    id_code_qualifier=validation('id_code_qualifier',split_tilde(cp2[8]))
                    clp_id_code=validation('clp_id_code',split_tilde(cp2[9]))
                except:
                    pass
                # validation_results = {
                # 'entity_type_qualifier': (entity_type_qualifier, split_tilde(cp2[2]),2),
                # 'pat_lastname': (pat_lastname, split_tilde(cp2[3]),3),
                # 'pat_firstname': (pat_firstname, split_tilde(cp2[4]),4),
                # 'pat_middlename': (pat_middlename, split_tilde(cp2[5]),5),
                # 'pat_prefix': (pat_prefix, split_tilde(cp2[6]),6),
                # 'pat_suffix': (pat_suffix, split_tilde(cp2[7]),7),
                # 'id_code_qualifier': (id_code_qualifier, split_tilde(cp2[8]),8),
                # 'clp_id_code': (clp_id_code, split_tilde(cp2[9]),9)
                # }


                validation_results = {}

                if entity_type_qualifier != 'notfound':
                    validation_results['entity_type_qualifier'] = (entity_type_qualifier, split_tilde(cp2[2]), 2)
                if pat_lastname != 'notfound':
                    validation_results['pat_lastname'] = (pat_lastname, split_tilde(cp2[3]), 3)
                if pat_firstname != 'notfound':
                    validation_results['pat_firstname'] = (pat_firstname, split_tilde(cp2[4]), 4)
                if pat_middlename != 'notfound':
                    validation_results['pat_middlename'] = (pat_middlename, split_tilde(cp2[5]), 5)
                if pat_prefix != 'notfound':
                    validation_results['pat_prefix'] = (pat_prefix, split_tilde(cp2[6]), 6)
                if pat_suffix != 'notfound':
                    validation_results['pat_suffix'] = (pat_suffix, split_tilde(cp2[7]), 7)
                if id_code_qualifier != 'notfound':
                    validation_results['id_code_qualifier'] = (id_code_qualifier, split_tilde(cp2[8]), 8)
                if clp_id_code != 'notfound':
                    validation_results['clp_id_code'] = (clp_id_code, split_tilde(cp2[9]), 9)              
                failed_entry=[
                    {'lines': clpline, 'value': value,'position':position,'field':'CLP','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
                
            elif clpdata.startswith('NM1*') and '*IL*' in clpdata:
                cp3=clpdata.split('*')
                sub_lastname,sub_firstname,sub_middlename='notfound','notfound','notfound'
                try:
                    sub_lastname=validation('pat_lastname',split_tilde(cp3[3]))
                    sub_firstname=validation('pat_firstname',split_tilde(cp3[4]))
                    sub_middlename=validation('pat_middlename',split_tilde(cp2[5]))
                except:
                    pass
                validation_results={}

                if sub_lastname != 'notfound':
                    validation_results['pat_lastname'] = (sub_lastname, split_tilde(cp3[3]), 3)
                if sub_firstname != 'notfound':
                    validation_results['pat_firstname'] = (sub_firstname, split_tilde(cp3[4]), 4)
                if sub_middlename != 'notfound':
                    validation_results['pat_middlename'] = (sub_middlename, split_tilde(cp3[5]), 5)
                # validation_results = {
                # 'sub_lastname': (sub_lastname, split_tilde(cp3[3]),3),
                # 'sub_firstname': (sub_firstname, split_tilde(cp3[4]),4),
                # 'sub_middlename': (sub_middlename, split_tilde(cp3[5]),5)
                # }
                failed_entry=[
                    {'lines': clpline, 'value': value,'position':position,'field':'CLP','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
                # print(hp1)
            elif clpdata.startswith('NM1*') and '*82*' in clpdata:
                cp4=clpdata.split('*')
                provider_lastname,provider_firstname,provider_middlename,provider_prefix,provider_suffix,provider_id_code_qualifier,provider_id_code='notfound','notfound','notfound','notfound','notfound','notfound','notfound'
                try:
                    provider_lastname=validation('provider_lastname',split_tilde(cp4[3]))
                    provider_firstname=validation('provider_firstname',split_tilde(cp4[4]))
                    provider_middlename=validation('provider_middlename',split_tilde(cp4[5]))
                    provider_prefix=validation('provider_prefix',split_tilde(cp4[6]))
                    provider_suffix=validation('provider_prefix',split_tilde(cp4[7]))
                    provider_id_code_qualifier=validation('provider_id_code_qualifier',split_tilde(cp4[8]))
                    provider_id_code=validation('provider_id_code',split_tilde(cp4[9]))
                except:
                    pass

                validation_results = {}

                if provider_lastname != 'notfound':
                    validation_results['provider_lastname'] = (provider_lastname, split_tilde(cp4[3]), 3)
                if provider_firstname != 'notfound':
                    validation_results['provider_firstname'] = (provider_firstname, split_tilde(cp4[4]), 4)
                if provider_middlename != 'notfound':
                    validation_results['provider_middlename'] = (provider_middlename, split_tilde(cp4[5]), 5)
                if provider_prefix != 'notfound':
                    validation_results['provider_prefix'] = (provider_prefix, split_tilde(cp4[6]), 6)
                if provider_suffix != 'notfound':
                    validation_results['provider_suffix'] = (provider_suffix, split_tilde(cp4[7]), 7)
                if provider_id_code_qualifier != 'notfound':
                    validation_results['provider_id_code_qualifier'] = (provider_id_code_qualifier, split_tilde(cp4[8]), 8)
                if provider_id_code != 'notfound':
                    validation_results['provider_id_code'] = (provider_id_code, split_tilde(cp4[9]), 9)

                failed_entry = [
                    {'lines': clpline, 'value': value, 'position': position, 'field': 'CLPData', 'title': var_name}
                    for var_name, (is_valid, value, position) in validation_results.items()
                    if not is_valid
                ]

                if failed_entry:
                    failed_info.append(failed_entry)

            clpline+=1
    svc_small=[]
    svc_lar=[]
    for index4,svc in enumerate(svc_chunks):
        svc_md=[]
        for index5,svc1 in enumerate(svc):
            if svc1.startswith('SVC*'):
                svc_fd=True
            elif index5!=len(svc)-1:
                if svc[index5+1].startswith('SVC*'):
                    svc_small.append(svc1)
                    svc_md.append(svc_small)
                    svc_fd=False
                    svc_small=[]
            elif index5==len(svc)-1:
                svc_small.append(svc1)
                svc_md.append(svc_small)
                svc_fd=False
                svc_small=[]
            if svc_fd:
                svc_small.append(svc1)
            if not svc_fd and len(svc_small)>0:
                svc_md.append(svc_small)
        svc_lar.append(svc_md)
    svc_line=header_count+len(clp_full[0])
    for index6,svcdata in enumerate(svc_lar):
        if index6>0:
            svc_line=svc_line+len(clp_full[index6])
        for index7,svcdata1 in enumerate(svcdata):
            for index8,svcdata2 in enumerate(svcdata1):
                if svcdata2.startswith('SVC*'):
                    sd1=svcdata2.split('*')
                    procedure_code,product_service_id,billed,paid_amount='notfound','notfound','notfound','notfound'
                    try:
                        procedurecode=sd1[1]
                        if ':' in procedurecode:
                            proc_len=len(procedurecode.split(':'))
                            if proc_len==1:
                                procedure_code=validation('procedure_code',split_tilde(procedurecode.split(':')[0]))
                                product_service_id=validation('product_service_id',split_tilde(procedurecode.split(':')[1]))
                        else:
                            pass
                        
                        billed=validation('service_charge_amt',split_tilde(sd1[2]))
                        paid_amount=validation('paid_amount',split_tilde(sd1[3]))
                    except:
                        pass
                    validation_results = {}

                    if procedure_code != 'notfound':
                        validation_results['procedure_code'] = (procedure_code, split_tilde(sd1[1].split(':')[0]), 1)
                    if product_service_id != 'notfound':
                        validation_results['product_service_id'] = (product_service_id, split_tilde(sd1[1].split(':')[1]), 1)
                    if billed != 'notfound':
                        validation_results['service_charge_amt'] = (billed, split_tilde(sd1[2]), 2)
                    if paid_amount != 'notfound':
                        validation_results['paid_amount'] = (paid_amount, split_tilde(sd1[3]), 3)

                    failed_entry = [
                        {'lines': svc_line, 'value': value, 'position': position, 'field': 'ServiceLine', 'title': var_name}
                        for var_name, (is_valid, value, position) in validation_results.items()
                        if not is_valid
                    ]

                    if failed_entry:
                        failed_info.append(failed_entry)

                    # validation_results = {
                    # 'procedure_code': (procedure_code, split_tilde(procedurecode),1),
                    # 'billed': (billed, split_tilde(sd1[2]),2),
                    # 'service_pay_amount': (paid_amount, split_tilde(sd1[3]),3)
                    # }
                    # failed_entry=[
                    #     {'lines': svc_line, 'value': value,'position':position,'field':'SVC','title':var_name}
                    #     for var_name, (is_valid, value,position) in validation_results.items()
                    #     if not is_valid
                    # ]
                    # if failed_entry:
                    #     failed_info.append(failed_entry)
                elif svcdata2.startswith('DTM*'):
                    sd3=svcdata2.split('*')
                    service_date='notfound'
                    service_date_qualifier='notfound'
                    try:
                        service_date_qualifier=validation('service_date_qualifier',split_tilde(sd3[1]))
                        service_date=validation('pay_date',split_tilde(sd3[2]))
                    except:
                        pass
                    validation_results = {}

                    if service_date != 'notfound':
                        validation_results['service_date'] = (service_date, split_tilde(sd3[2]), 2)
                    if service_date_qualifier != 'notfound':
                        validation_results['service_date_qualifier'] = (service_date_qualifier, split_tilde(sd3[1]), 1)
                    

                    failed_entry = [
                        {'lines': svc_line, 'value': value, 'position': position, 'field': 'ServiceLine', 'title': var_name}
                        for var_name, (is_valid, value, position) in validation_results.items()
                        if not is_valid
                    ]

                    if failed_entry:
                        failed_info.append(failed_entry)

                elif svcdata2.startswith('CAS*'):
                    sd2=svcdata2.split('*')
                    claim_grp_code,adjustment_reason_code,adjustment_amt='notfound','notfound','notfound'
                    try:
                        claim_grp_code=validation('claim_grp_code',split_tilde(sd2[1]))
                        adjustment_reason_code=validation('adjustment_reason_code',split_tilde(sd2[2]))
                        adjustment_amt=validation('adjustment_amt',split_tilde(sd2[3]))
                    except:
                        pass

                    validation_results = {}

                    if claim_grp_code != 'notfound':
                        validation_results['claim_grp_code'] = (claim_grp_code, split_tilde(sd2[1]), 1)
                    if adjustment_reason_code != 'notfound':
                        validation_results['adjustment_reason_code'] = (adjustment_reason_code, split_tilde(sd2[2]), 2)
                    if adjustment_amt != 'notfound':
                        validation_results['adjustment_amt'] = (adjustment_amt, split_tilde(sd2[3]), 3)

                    failed_entry = [
                        {'lines': svc_line, 'value': value, 'position': position, 'field': 'ServiceLine', 'title': var_name}
                        for var_name, (is_valid, value, position) in validation_results.items()
                        if not is_valid
                    ]

                    if failed_entry:
                        failed_info.append(failed_entry)
                # elif svcdata2.startswith('AMT*'):
            

                svc_line+=1
    summary_lines=[]
    summary_fd=False
    for index9,summ in enumerate(full_lines):
        if index9==totalcount2:
            summary_fd=True
        if summary_fd:
            summary_lines.append(full_lines[index9])
    summary_lines_count=totalcount2
    for index10,summary in enumerate(summary_lines):
        if summary.startswith('SE*'):
            sud1=summary.split('*')
            included_sgmt_no,trans_set_control_number='notfound','notfound'
            try:
                included_sgmt_no=validation('included_sgmt_no',split_tilde(sud1[1]))
                trans_set_control_number=validation('trans_set_control_number',split_tilde(sud1[2]))
            except:
                pass
            validation_results = {}

            if included_sgmt_no != 'notfound':
                validation_results['included_sgmt_no'] = (included_sgmt_no, split_tilde(sud1[1]), 1)
            if trans_set_control_number != 'notfound':
                validation_results['trans_set_control_number'] = (trans_set_control_number, split_tilde(sud1[2]), 2)
            failed_entry = [
                {'lines': summary_lines_count, 'value': value, 'position': position, 'field': 'SummaryLine', 'title': var_name}
                for var_name, (is_valid, value, position) in validation_results.items()
                if not is_valid
            ]

            if failed_entry:
                failed_info.append(failed_entry)
        elif summary.startswith('GE*'):
            sud2=summary.split('*')
            number_of_transaction_set,group_control_no='notfound','notfound'
            try:
                number_of_transaction_set=validation('number_of_transaction_set',split_tilde(sud2[1]))
                group_control_no=validation('group_control_no',split_tilde(sud2[2]))
            except:
                pass
            validation_results = {}
            if number_of_transaction_set != 'notfound':
                validation_results['number_of_transaction_set'] = (number_of_transaction_set, split_tilde(sud2[1]), 1)
            if group_control_no != 'notfound':
                validation_results['group_control_no'] = (group_control_no, split_tilde(sud2[2]), 2)
            failed_entry = [
                {'lines': summary_lines_count, 'value': value, 'position': position, 'field': 'SummaryLine', 'title': var_name}
                for var_name, (is_valid, value, position) in validation_results.items()
                if not is_valid
            ]

            if failed_entry:
                failed_info.append(failed_entry)
        elif summary.startswith('IEA*'):
            no_of_included_segments,interchange_control_no='notfound','notfound'
            sud3=summary.split('*')
            try:
                no_of_included_segments=validation('no_of_included_segments',split_tilde(sud3[1]))
                interchange_control_no=validation('interchange_control_no',split_tilde(sud3[2]))
            except:
                pass
            validation_results={}
            if no_of_included_segments != 'notfound':
                validation_results['no_of_included_segments'] = (no_of_included_segments, split_tilde(sud3[1]), 1)
            if interchange_control_no != 'notfound':
                validation_results['interchange_control_no'] = (interchange_control_no, split_tilde(sud3[2]), 2)
            failed_entry = [
                {'lines': summary_lines_count, 'value': value, 'position': position, 'field': 'SummaryLine', 'title': var_name}
                for var_name, (is_valid, value, position) in validation_results.items()
                if not is_valid
            ]
            if failed_entry:
                failed_info.append(failed_entry)
        summary_lines_count+=1

        # pass

    # print(failed_info)
    # os.remove(new_path)
    return failed_info


def process_files(file_data, excel_path):
    # Initialize a list to store all rows for the DataFrame
    rows = []

    for file_info in file_data:
        file_name = file_info[0].get('file_name', 'Unknown')

        # Initialize sections
        header_data = ""
        clp_data = ""
        service_line_data = ""
        summary_data = ""

        for item in file_info[1:]:
            for entry in item:
                field = entry.get('field', '')
                value = entry.get('value', '')
                title = entry.get('title', '')

                # Format and concatenate by section
                line = f'{title}: {value}'
                if field == 'Header':
                    header_data += (line + '\n') if header_data else line
                elif field == 'CLP':
                    clp_data += (line + '\n') if clp_data else line
                elif field == 'ServiceLine':
                    service_line_data += (line + '\n') if service_line_data else line
                elif field == 'Summary':
                    summary_data += (line + '\n') if summary_data else line

        # Append one row per file, replace empty strings with "VALIDATED"
        rows.append({
            'File Name': file_name,
            'Headers': header_data.strip() if header_data.strip() else "VALIDATED",
            'CLP Line': clp_data.strip() if clp_data.strip() else "VALIDATED",
            'Service Line': service_line_data.strip() if service_line_data.strip() else "VALIDATED",
            'Summary Line': summary_data.strip() if summary_data.strip() else "VALIDATED"
        })

    # Create DataFrame with renamed headers
    df = pd.DataFrame(rows)

    # Write to Excel using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "835 Validate"

    # Write column headers
    headers = df.columns.tolist()
    ws.append(headers)

    # Write data rows
    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Apply formatting
    # Define styles
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray background
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Apply styles to headers
    for cell in ws[1]:  # Row 1 contains headers
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border

    # Apply borders to all data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Adjust column widths (optional, for better readability)
    column_widths = []
    for row in ws.rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths.append(len(str(cell.value)))
    for i, column_width in enumerate(column_widths, 1):  # +1 because column index starts at 1
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = column_width + 2

    wb.save(excel_path)

def validation_multiplefile(folderpath,excel_path):
    print('validation_multiplefile')
    def split_tilde(data):
        if '~' in data:
            data=data.split('~')[0]
            return data
        elif '~' not in data:
            return data
    def validation(title,data):
        if title in ['sender_id','rec_id','payer_id']:
            if re.fullmatch(r'^[a-zA-Z0-9]{1,15}$', data): #aplha numeric max 1 - 15 
                return True
            return False
        elif title in ['auth_info_qualifier']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9]{2}$", data))
        elif title in ['authorization_info']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9 ]{10}$", data))
        elif title in ['interchange_sender_id']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9 ]{15}$", data))
        elif title in ['interchange_date']:
            return bool(re.fullmatch(r"^\d{2}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$", data))
        elif title in ['interchange_time']:
            return bool(re.fullmatch(r"^([01][0-9]|2[0-3])[0-5][0-9]$", data))
        elif title in ['interchange_standard_id']:
            return bool(re.fullmatch(r"^.{1}$", data))
        elif title in ['transcation_heading_code']:
            if data in 'CDHIPUX':
                return True
            else:
                return False
        elif title in ['pay_format_code']:
            if data in ['CCP','']:
                return True
            else:
                return False
        elif title in ['ref_identification_qul']:
            if data =='EV':
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,2}$", data))
        elif title in ['id_no_qualifier']:
            if data in ['01','']:
                return True
            else:
                return False
            
        elif title in ['identification_bpr']:
            if data in ['999999999','']:
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9]{3,12}$", data))
        
        elif title in ['account_number']:
            if data in ['']:
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,35}$", data))            
        elif title in ['effective_entry_date']:
            pattern = r"^(19|20)\d{2}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['included_sgmt_no']:
            return bool(re.fullmatch(r"^\d{1,10}$", data))
        elif title in ['number_of_transaction_set']:
            return bool(re.fullmatch(r"\d{1,6}", data))
        elif title in ['group_control_no']:
            return bool(re.fullmatch(r"\d{1,9}", data))
        elif title in ['no_of_included_segments']:
            return bool(re.fullmatch(r"\d{1,5}", data))
        elif title in ['interchange_control_no']:
            return bool(re.fullmatch(r"\d{9}", data))
        elif title in ['trans_set_control_number']:
            return bool(re.fullmatch(r"^\d{4,9}$", data))
        elif title in ['acc_no_qualifier_code']:
            if data in ['DA','']:
                return True
            else:
                return False
        elif title in ['originating_company_identfier']:
            if data in ['9999999999','']:
                return True
            else:
                pattern = r"^[a-zA-Z0-9]{10}$"
                return bool(re.fullmatch(pattern, data))
                
        elif title in ['originating_supplemental_code']:
            if data in ['','199999999']:
                return True
            else:
                pattern = r"^[a-zA-Z0-9]{9}$"
                return bool(re.fullmatch(pattern, data))
            
        elif title in ['rdfi_id_number']:
            if data in ['','01']:
                return True
            else:
                return False
        elif title in ['rdfi_identification_number']:
            if data in ['','999999999']:
                return True
            else:
                pattern = r"^[a-zA-Z0-9]{3,12}$"
                return bool(re.fullmatch(pattern, data))
        
        elif title in ['cre_deb_flag_code']:
            if data in ['C','D']:
                return True
            else:
                return False
        elif title in ['intchangecontrolno']:
            return bool(re.fullmatch(r"\d{9}", data)) #numeric 9 digit
        elif title in ['grp_con_no','tans_set_con_no']:
            return bool(re.fullmatch(r'^\d{1,9}$', data)) #numeric max 9 digit
        elif title in ['total_pay_amt']:
            pattern = r'^-?\d{1,18}(\.\d{1,2})?$'
            return bool(re.fullmatch(pattern, data))
        elif title in ['pay_method_code']:
            if data in ['ACH', 'CHK', 'NON']:
                return True
            else:
                return False
        elif title in ['bank_acc_no']:
            return bool(re.fullmatch(r'^[a-zA-Z0-9]{1,17}$', data))
        elif title in ['check_eft_trace_no']:
            return bool(re.fullmatch(r'^[a-zA-Z0-9]{1,20}$', data))
        elif title in ['payer_city','payee_city']:
            return bool(re.fullmatch(r'^[a-zA-Z0-9 ]{1,30}$', data))
        elif title in ['trace_type_code']:
            if data =='1':
                return True
            else:
                return False
        elif title in ['check_no']:
            return bool(re.fullmatch(r'^[a-zA-Z0-9]{1,30}$', data))
        elif title in ['payer_id','payee_id']:
            # Validates if the input is alphanumeric and at most 15 characters long.
            return bool(re.fullmatch(r'^[a-zA-Z0-9]{1,15}$', data))
        elif title in ['curr_code']:
            if data in ['USD','CAD']:
                return True
            else:
                return False
        elif title in ['pay_ref_qul']:
            if len(data)==2 and data in ['EV']:
                return True
            else:
                return False
        elif title in ['ref_identification_num']:
            return bool(re.fullmatch(r"^.{1,30}$", data))
        elif title in ['date_aualifier']:
            if data in ['405']:
                return True
            else:
                return False
        elif title in ['service_date_qualifier']:
            if title in ['472','150','151']:
                return True
            else:
                return bool(re.fullmatch(r"^\d{3}$", data))
        elif title in ['pay_date']:
            # Validates if the input is an 8-digit date in YYYYMMDD format.
            pattern = r"^(19|20)\d{2}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['payer_name','payee_name']:
            # Validates if the input is alphanumeric and at most 60 characters long.
            return bool(re.fullmatch(r'^[a-zA-Z0-9 .\-]{1,60}$', data))
        elif title in ['payee_id_code_qualifier']:
            if data in ['FI','XX']:
                return True
            else:
                return False
        elif title in ['payer_address','payee_address']:
            # Validates if the input is alphanumeric and at most 55 characters long.
            return bool(re.fullmatch(r'^[a-zA-Z0-9 \-]{1,55}$', data))
        elif title in ['payer_state','payee_state']:
            # Validates if the input is exactly 2 uppercase letters (US State Code).
            return bool(re.fullmatch(r'^[A-Z]{2}$', data))
        elif title in ['payer_zip','payee_zip']:
            # Validates if the input is a 5-digit or 9-digit ZIP code (ZIP+4 format).
            return bool(re.fullmatch(r'^\d{5}(\d{4})?$', data))
        elif title in ['service_line_no']:
            # Validates if the input is a numeric value with 1 to 9 digits.
            return bool(re.fullmatch(r'^\d{1,9}$', data))
        elif title in ['claim_no']:
            # Validates if the input is alphanumeric with a maximum of 38 characters.
            return bool(re.fullmatch(r'^[a-zA-Z0-9]{1,38}$', data))
        elif title in ['claim_sts_code']:
            # Validates if the input is a numeric value with 1 to 4 digits.
            return bool(re.fullmatch(r'^\d{1,2}$', data))
        elif title in ['total_charges','total_pay_amt']:
            # Decimal, max 18 digits, 2 decimal places
            return bool(re.fullmatch(r'^-?\d{1,18}(\.\d{1,2})?$', data))
        elif title in ['clm_indicator_code']:
            pattern = r"^[a-zA-Z0-9]{1,2}$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['payer_claim_control_no']:
            pattern = r"^[a-zA-Z0-9]{1,30}$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['pat_lastname']:
            return bool(re.fullmatch(r'^[a-zA-Z ]{1,35}$', data))
        elif title in ['entity_type_qualifier']:
            if data == '1':
                return True
            else:
                return False
        elif title in ['pat_firstname']:
            return bool(re.fullmatch(r'^[a-zA-Z ]{1,25}$', data))
        elif title in ['pat_middlename']:
            if data=='':
                return True
            else:
                pattern = r"^[a-zA-Z0-9]{1,25}$"
                return bool(re.fullmatch(pattern, data))
        elif title in ['pat_suffix','pat_prefix']:
            if data == '':
                return True
            pattern = r"^[a-zA-Z0-9]{1,10}$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['provider_lastname']:
            if data == '':
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9 ]{1,60}$", data))
        elif title in ['provider_firstname']:
            if data == '':
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9 ]{1,35}$", data))
        elif title in ['provider_middlename']:
            if data == '':
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,25}$", data))
        elif title in ['provider_prefix']:
            if data == '':
                return True
            else:
                return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,10}$", data))
        elif title in ['provider_id_code_qualifier']:
            if data in ['XX','FI']:
                return True
            else:
                return False
        elif title in ['provider_id_code']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9]{2,80}$", data))
        elif title in ['clp_id_code']:
            pattern = r"^[a-zA-Z0-9]{2,80}$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['id_code_qualifier']:
            pattern = r"^[a-zA-Z0-9]{1,2}$"
            return bool(re.fullmatch(pattern, data))
        elif title in ['procedure_code']:
            if data in ['HC','NU']:
                return True
            else:
                return False
        elif title in ['product_service_id']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,48}$", data))
            
        elif title in ['service_charge_amt','service_pay_amount','paid_amount']:
            if data in  ['0','']:
                return True
            else:
                return bool(re.fullmatch(r"^-?\d{1,18}(\.\d{1,2})?$", data))

        elif title in ['claim_grp_code']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,2}$", data))
        elif title in ['adjustment_reason_code']:
            return bool(re.fullmatch(r"^[a-zA-Z0-9]{1,5}$", data))
        elif title in ['adjustment_amt']:
            return bool(re.fullmatch(r"^-?\d{1,18}(\.\d{1,2})?$", data))

        
    

    def chunk_file_content(file_content):
        # Split the file content into lines
        lines = file_content.split('\n')
        
        # Identify lines that start with 'CLP*'
        clp_lines = [i for i, line in enumerate(lines) if line.startswith("CLP*")]
        
        # Identify the header data (lines before the first CLP*)
        header_data = lines[:clp_lines[0]] if clp_lines else []

        # Initialize a list for each CLP data chunk
        clp_data_chunks = []
        
        # Loop through each CLP, and chunk data accordingly
        for i in range(len(clp_lines)):
            start = clp_lines[i]
            if i + 1 < len(clp_lines):
                end = clp_lines[i + 1]
            else:
                # For the last CLP, we want to go up to the first IEA* line
                iea_line = next((j for j, line in enumerate(lines) if line.startswith("IEA*")), len(lines))
                end = iea_line
            
            clp_data_chunks.append((f"CLP_{i + 1}", lines[start:end]))
        clpmaindata=[]
        for indexx,clpd in enumerate(clp_data_chunks):
            clpmaindata.append(clpd[1])
        # clpdata=clp_data_chunks[1]
        svc_final_chunk=[]
        for indexxx,clpdata in enumerate(clpmaindata):
            svc_chunk=[]
            svc_fd=False
            
            for indx4,svcchunk in enumerate(clpdata):
                if 'SVC*' in svcchunk:
                    svc_fd=True
                if svcchunk.startswith("SE*"):
                    break
                if svc_fd:
                    svc_chunk.append(svcchunk)
            svc_final_chunk.append(svc_chunk)
        clp_data_final=[]
        for clpdata in clpmaindata:
            clp_data=[]
            for clpdata2 in clpdata:
                if clpdata2.startswith('SVC*'):
                    break
                else:
                    clp_data.append(clpdata2)
            clp_data_final.append(clp_data)
        return header_data, clp_data_chunks,svc_final_chunk,lines,clp_data_final
    files=os.listdir(folderpath)
    files=[file for file in files if file.endswith('.835')]
    final_failed_info=[]
    for file_name in files:
        failed_info=[]
        failed_info.append({'file_name': file_name})
        file_path=os.path.join(folderpath, file_name)
        # file_path2=os.path.dirname(file_path)
        # new_path=os.path.join(file_path2,'temp.835')
        # shutil.copy(file_path,new_path)
        with open(file_path, "r", encoding="utf-8") as f1:
                file_content = f1.read()
                # f1.close()
        header_data, clp_chunks,svc_chunks,full_lines,clp_full = chunk_file_content(file_content)
        total_lines_count=len(full_lines)
        header_count=len(header_data)
        clp_count=0
        clpcount2=0
        for clpfu in clp_full:
            clpcount2+=len(clpfu)
        svc_count2=0
        for svcfu in svc_chunks:
            svc_count2+=len(svcfu)
        totalcount2=header_count+clpcount2+svc_count2
        for clpd1 in clp_full:
            clp_count+=len(clpd1)
        svc_count=0
        for svcd1 in svc_chunks:
            svc_count+=len(svcd1)
        for index1,header in enumerate(header_data):
            if header.startswith('ISA*'):
                hp1=header.split('*')
                # for ind_,hp1_t in enumerate(hp1):
                #     hp1[ind_]=hp1_t.strip()
                auth_info_qualifier=validation('auth_info_qualifier',split_tilde(hp1[1]))
                authorization_info=validation('authorization_info',split_tilde(hp1[2]))
                security_info_qualifier=validation('auth_info_qualifier',split_tilde(hp1[3]))
                security_information=validation('authorization_info',split_tilde(hp1[4]))
                interchange_sender_id=validation('interchange_sender_id',split_tilde(hp1[6]))
                interchange_id_qualifier=validation('auth_info_qualifier',split_tilde(hp1[7]))
                interchange_receiver_id=validation('interchange_sender_id',split_tilde(hp1[8]))
                interchange_date=validation('interchange_date',split_tilde(hp1[9]))
                interchange_time=validation('interchange_time',split_tilde(hp1[10]))
                interchange_standard_id=validation('interchange_standard_id',split_tilde(hp1[11]))

                # sender_id_validate=validation('sender_id',hp1[6])
                # rec_id_validate=validation('rec_id',hp1[8])
                interchange_controlno_validate=validation('intchangecontrolno',hp1[13])
                validation_results = {
                    'auth_info_qualifier': (auth_info_qualifier, split_tilde(hp1[1]), 1),
                    'authorization_info': (authorization_info, split_tilde(hp1[2]), 2),
                    'security_info_qualifier': (security_info_qualifier, split_tilde(hp1[3]), 3),
                    'security_information': (security_information, split_tilde(hp1[4]), 4),
                    'interchange_sender_id': (interchange_sender_id, split_tilde(hp1[6]), 6),
                    # 'sender_id': (sender_id_validate, hp1[6], 6),
                    'interchange_id_qualifier': (interchange_id_qualifier, split_tilde(hp1[7]), 7),
                    'interchange_receiver_id': (interchange_receiver_id, split_tilde(hp1[8]), 8),
                    'interchange_date': (interchange_date, split_tilde(hp1[9]), 9),
                    'interchange_time': (interchange_time, split_tilde(hp1[10]), 10),
                    'interchange_standard_id': (interchange_standard_id, split_tilde(hp1[11]), 11),
                    'intchangecontrolno': (interchange_controlno_validate, hp1[13], 13),
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('GS*'):
                hp2=header.split('*')
                sender_id2_vaidate=validation('sender_id',hp2[2])
                rec_id2_validate=validation('rec_id',hp2[3])
                grp_control_number_validate=validation('grp_con_no',hp2[6])
                validation_results = {
                    'sender_id2': (sender_id2_vaidate, hp2[2],2),
                    'rec_id2': (rec_id2_validate, hp2[3],3),
                    'grp_control_number': (grp_control_number_validate, hp2[6],6)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('ST*835'):
                hp3=header.split('*')
                trans_set_con_no_validate=validation('tans_set_con_no',split_tilde(hp3[2]))
                validation_results = {
                    'trans_set_con_no': (trans_set_con_no_validate, split_tilde(hp3[2]),2)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('BPR*'):
                hp4=header.split('*')
                transcation_heading_code=validation('transcation_heading_code',split_tilde(hp4[1]))
                total_pay_amt=validation('total_pay_amt',hp4[2])
                cre_deb_flag_code=validation('cre_deb_flag_code',split_tilde(hp4[3]))
                pay_method_code=validation('pay_method_code',split_tilde(hp4[4]))
                pay_format_code=validation('pay_format_code',split_tilde(hp4[5]))
                id_no_qualifier=validation('id_no_qualifier',split_tilde(hp4[6]))
                identification_bpr=validation('identification_bpr',split_tilde(hp4[7]))
                acc_no_qualifier_code_bpr1=validation('acc_no_qualifier_code',split_tilde(hp4[8]))
                account_number_bpr1=validation('account_number',split_tilde(hp4[9]))
                originating_company_identfier=validation('originating_company_identfier',split_tilde(hp4[10]))
                originating_supplemental_code=validation('originating_supplemental_code',split_tilde(hp4[11]))
                rdfi_id_number=validation('rdfi_id_number',split_tilde(hp4[12]))
                rdfi_identification_number=validation('rdfi_identification_number',split_tilde(hp4[13]))
                account_no_qual_code_bpr2=validation('acc_no_qualifier_code',split_tilde(hp4[14]))
                account_number_bpr2=validation('account_number',split_tilde(hp4[15]))
                effective_entry_date=validation('effective_entry_date',split_tilde(hp4[16]))
                
                validation_results = {
                    'transcation_heading_code': (transcation_heading_code, split_tilde(hp4[1]), 1),
                    'total_pay_amt': (total_pay_amt, hp4[2], 2),
                    'cre_deb_flag_code': (cre_deb_flag_code, split_tilde(hp4[3]), 3),
                    'pay_method_code': (pay_method_code, split_tilde(hp4[4]), 4),
                    'pay_format_code': (pay_format_code, split_tilde(hp4[5]), 5),
                    'id_no_qualifier': (id_no_qualifier, split_tilde(hp4[6]), 6),
                    'identification_bpr': (identification_bpr, split_tilde(hp4[7]), 7),
                    'acc_no_qualifier_code_bpr1': (acc_no_qualifier_code_bpr1, split_tilde(hp4[8]), 8),
                    'account_number_bpr1': (account_number_bpr1, split_tilde(hp4[9]), 9),
                    'originating_company_identfier': (originating_company_identfier, split_tilde(hp4[10]), 10),
                    'originating_supplemental_code': (originating_supplemental_code, split_tilde(hp4[11]), 11),
                    'rdfi_id_number': (rdfi_id_number, split_tilde(hp4[12]), 12),
                    'rdfi_identification_number': (rdfi_identification_number, split_tilde(hp4[13]), 13),
                    'account_no_qual_code_bpr2': (account_no_qual_code_bpr2, split_tilde(hp4[14]), 14),
                    'account_number_bpr2': (account_number_bpr2, split_tilde(hp4[15]), 15),
                    'effective_entry_date': (effective_entry_date, split_tilde(hp4[16]), 16)
                }

                failed_entry = [
                    {'lines': index1, 'value': value, 'position': position, 'field': 'Header', 'title': var_name}
                    for var_name, (is_valid, value, position) in validation_results.items()
                    if not is_valid
                ]

                if failed_entry:
                    failed_info.append(failed_entry)



            elif header.startswith('TRN*'):
                hp5=header.split('*')
                trace_type_code_validate=validation('trace_type_code',hp5[1])
                check_no_validate=validation('check_no',hp5[2])
                payer_id_validate=validation('payer_id',split_tilde(hp5[3]))
                validation_results = {
                    'trace_type_code': (trace_type_code_validate, hp5[1],1),
                    'check_no': (check_no_validate, hp5[2],2),
                    'payer_id': (payer_id_validate, split_tilde(hp5[3]),3)

                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('CUR*'):
                hp6=header.split('*')
                curr_code_validate=validation('curr_code',split_tilde(hp6[1]))
                validation_results = {
                    'curr_code': (curr_code_validate, split_tilde(hp6[1]),1)

                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('REF*') and header_data[index1+1].startswith('LX*'):
                hp16=header.split('*')
                payee_id=validation('payee_id',split_tilde(hp16[2]))
                validation_results = {
                    'payee_id': (payee_id, split_tilde(hp16[2]),2)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('LX*'):
                hp17=header.split('*')
                service_line_no=validation('service_line_no',split_tilde(hp17[1]))
                validation_results = {
                    'service_line_no': (service_line_no, split_tilde(hp17[1]),1)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('REF*') and (header_data[index1+2].startswith('N1*PR*') or header_data[index1+1].startswith('N1*PR*')):
                hp7=header.split('*')
                ref_identification_qul_validate=validation('pay_ref_qul',split_tilde(hp7[1]))
                ref_identification_num=validation('ref_identification_num',split_tilde(hp7[2]))
                validation_results = {
                    'ref_identification_qul': (ref_identification_qul_validate, split_tilde(hp7[1]),1),
                    'ref_identification_num': (ref_identification_num, split_tilde(hp7[2]),2)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('DTM*'):
                hp8=header.split('*')
                date_qaualifier=validation('date_aualifier',split_tilde(hp8[1]))
                pay_date=validation('pay_date',split_tilde(hp8[2]))
                validation_results = {
                    'date_qaualifier':(date_qaualifier, split_tilde(hp8[1]),1),
                    'pay_date': (pay_date, split_tilde(hp8[2]),2)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('N1*PR*'):
                hp9=header.split('*')
                payer_name=validation('payer_name',split_tilde(hp9[2]))
                validation_results = {
                    'payer_name': (payer_name, split_tilde(hp9[2]),2)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('N3*') and (header_data[index1-2].startswith('N1*PR*') or header_data[index1-1].startswith('N1*PR*')):
                hp10=header.split('*')
                payer_address=validation('payer_address',split_tilde(hp10[1]))
                validation_results = {
                    'payer_address': (payer_address, split_tilde(hp10[1]),1)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('N4*') and (header_data[index1-2].startswith('N1*PR*') or header_data[index1-1].startswith('N1*PR*')):
                hp11=header.split('*')
                payer_city=validation('payer_city',split_tilde(hp11[1]))
                payer_state=validation('payer_state',split_tilde(hp11[2]))
                payer_zip=validation('payer_zip',split_tilde(hp11[3]))
                validation_results = {
                    'payer_city': (payer_city, split_tilde(hp11[1]),1),
                    'payer_state': (payer_state, split_tilde(hp11[2]),2),
                    'payer_zip': (payer_zip, split_tilde(hp11[3]),3)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('REF*') and (header_data[index1+2].startswith('N1*PE*') or header_data[index1+1].startswith('N1*PE*')):
                hp12=header.split('*')
                payer_id=validation('payer_id',split_tilde(hp12[2]))
                validation_results = {
                    'payer_id': (payer_id, split_tilde(hp12[2]),2)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('N1*PE*'):
                hp13=header.split('*')
                payee_name=validation('payee_name',split_tilde(hp13[2]))
                payee_id_code_qualifier=validation('payee_id_code_qualifier',split_tilde(hp13[3]))
                validation_results = {
                    'payee_name': (payee_name, split_tilde(hp13[2]),2),
                    'payee_id_code_qualifier': (payee_id_code_qualifier, split_tilde(hp13[3]),3)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('N3*') and (header_data[index1-2].startswith('N1*PE*') or header_data[index1-1].startswith('N1*PE*')):
                hp14=header.split('*')
                payee_address=validation('payee_address',split_tilde(hp14[1]))
                validation_results = {
                    'payee_address': (payee_address, split_tilde(hp13[2]),2)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            elif header.startswith('N4*') and (header_data[index1-2].startswith('N1*PE') or header_data[index1-1].startswith('N1*PE*')):
                hp15=header.split('*')
                payee_city=validation('payee_city',split_tilde(hp15[1]))
                payee_state=validation('payee_state',split_tilde(hp15[2]))
                payee_zip=validation('payee_zip',split_tilde(hp15[3]))
                validation_results = {
                    'payee_city': (payee_city, split_tilde(hp15[1]),1),
                    'payee_state': (payee_state, split_tilde(hp15[2]),2),
                    'payee_zip': (payee_zip, split_tilde(hp15[3]),3)
                }
                failed_entry=[
                    {'lines': index1, 'value': value,'position':position,'field':'Header','title':var_name}
                    for var_name, (is_valid, value,position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            
        for index2,clpd2 in enumerate(clp_full): 
            if index2==0:
                clpline=len(header_data)
            elif index2==1:
                clpline=len(header_data)+len(clp_full[index2-1])+len(svc_chunks[index2-1])
            else:
                clpline=clpline+len(svc_chunks[index2-1])
            for index3,clpdata in enumerate(clpd2):
                if clpdata.startswith('CLP*'):
                    cp1=clpdata.split('*')
                    claim_no=validation('claim_no',split_tilde(cp1[1]))
                    claim_sts_code=validation('claim_sts_code',split_tilde(cp1[2]))
                    total_charges=validation('total_charges',split_tilde(cp1[3]))
                    total_pay_amt=validation('total_pay_amt',split_tilde(cp1[4]))
                    pat_responsibility_amt=validation('total_pay_amt',split_tilde(cp1[5]))
                    clm_indicator_code=validation('clm_indicator_code',split_tilde(cp1[6]))
                    payer_claim_control_no=validation('payer_claim_control_no',split_tilde(cp1[7]))

                    validation_results = {
                    'claim_no': (claim_no, split_tilde(cp1[1]),1),
                    'claim_sts_code': (claim_sts_code, split_tilde(cp1[2]),2),
                    'total_charges': (total_charges, split_tilde(cp1[3]),3),
                    'total_pay_amt': (total_pay_amt, split_tilde(cp1[4]),4),
                    'pat_responsibility_amt': (pat_responsibility_amt, split_tilde(cp1[5]),5),
                    'clm_indicator_code': (clm_indicator_code, split_tilde(cp1[6]),6),
                    'payer_claim_control_no': (payer_claim_control_no, split_tilde(cp1[7]),7)
                    }
                    failed_entry=[
                        {'lines': clpline, 'value': value,'position':position,'field':'CLP','title':var_name}
                        for var_name, (is_valid, value,position) in validation_results.items()
                        if not is_valid
                    ]
                    if failed_entry:
                        failed_info.append(failed_entry)
                    
                elif clpdata.startswith('NM1*') and '*QC*' in clpdata:
                    cp2=clpdata.split('*')
                    entity_type_qualifier,pat_firstname,pat_lastname,pat_middlename,pat_prefix,pat_suffix,id_code_qualifier,clp_id_code='notfound','notfound','notfound','notfound','notfound','notfound','notfound','notfound'
                    try:
                        entity_type_qualifier=validation('entity_type_qualifier',split_tilde(cp2[2]))
                        pat_lastname=validation('pat_lastname',split_tilde(cp2[3]))
                        pat_firstname=validation('pat_firstname',split_tilde(cp2[4]))
                        pat_middlename=validation('pat_middlename',split_tilde(cp2[5]))
                        pat_prefix=validation('pat_prefix',split_tilde(cp2[6]))
                        pat_suffix=validation('pat_suffix',split_tilde(cp2[7]))
                        id_code_qualifier=validation('id_code_qualifier',split_tilde(cp2[8]))
                        clp_id_code=validation('clp_id_code',split_tilde(cp2[9]))
                    except:
                        pass
                    # validation_results = {
                    # 'entity_type_qualifier': (entity_type_qualifier, split_tilde(cp2[2]),2),
                    # 'pat_lastname': (pat_lastname, split_tilde(cp2[3]),3),
                    # 'pat_firstname': (pat_firstname, split_tilde(cp2[4]),4),
                    # 'pat_middlename': (pat_middlename, split_tilde(cp2[5]),5),
                    # 'pat_prefix': (pat_prefix, split_tilde(cp2[6]),6),
                    # 'pat_suffix': (pat_suffix, split_tilde(cp2[7]),7),
                    # 'id_code_qualifier': (id_code_qualifier, split_tilde(cp2[8]),8),
                    # 'clp_id_code': (clp_id_code, split_tilde(cp2[9]),9)
                    # }


                    validation_results = {}

                    if entity_type_qualifier != 'notfound':
                        validation_results['entity_type_qualifier'] = (entity_type_qualifier, split_tilde(cp2[2]), 2)
                    if pat_lastname != 'notfound':
                        validation_results['pat_lastname'] = (pat_lastname, split_tilde(cp2[3]), 3)
                    if pat_firstname != 'notfound':
                        validation_results['pat_firstname'] = (pat_firstname, split_tilde(cp2[4]), 4)
                    if pat_middlename != 'notfound':
                        validation_results['pat_middlename'] = (pat_middlename, split_tilde(cp2[5]), 5)
                    if pat_prefix != 'notfound':
                        validation_results['pat_prefix'] = (pat_prefix, split_tilde(cp2[6]), 6)
                    if pat_suffix != 'notfound':
                        validation_results['pat_suffix'] = (pat_suffix, split_tilde(cp2[7]), 7)
                    if id_code_qualifier != 'notfound':
                        validation_results['id_code_qualifier'] = (id_code_qualifier, split_tilde(cp2[8]), 8)
                    if clp_id_code != 'notfound':
                        validation_results['clp_id_code'] = (clp_id_code, split_tilde(cp2[9]), 9)              
                    failed_entry=[
                        {'lines': clpline, 'value': value,'position':position,'field':'CLP','title':var_name}
                        for var_name, (is_valid, value,position) in validation_results.items()
                        if not is_valid
                    ]
                    if failed_entry:
                        failed_info.append(failed_entry)
                    
                elif clpdata.startswith('NM1*') and '*IL*' in clpdata:
                    cp3=clpdata.split('*')
                    sub_lastname,sub_firstname,sub_middlename='notfound','notfound','notfound'
                    try:
                        sub_lastname=validation('pat_lastname',split_tilde(cp3[3]))
                        sub_firstname=validation('pat_firstname',split_tilde(cp3[4]))
                        sub_middlename=validation('pat_middlename',split_tilde(cp2[5]))
                    except:
                        pass
                    validation_results={}

                    if sub_lastname != 'notfound':
                        validation_results['pat_lastname'] = (sub_lastname, split_tilde(cp3[3]), 3)
                    if sub_firstname != 'notfound':
                        validation_results['pat_firstname'] = (sub_firstname, split_tilde(cp3[4]), 4)
                    if sub_middlename != 'notfound':
                        validation_results['pat_middlename'] = (sub_middlename, split_tilde(cp3[5]), 5)
                    # validation_results = {
                    # 'sub_lastname': (sub_lastname, split_tilde(cp3[3]),3),
                    # 'sub_firstname': (sub_firstname, split_tilde(cp3[4]),4),
                    # 'sub_middlename': (sub_middlename, split_tilde(cp3[5]),5)
                    # }
                    failed_entry=[
                        {'lines': clpline, 'value': value,'position':position,'field':'CLP','title':var_name}
                        for var_name, (is_valid, value,position) in validation_results.items()
                        if not is_valid
                    ]
                    if failed_entry:
                        failed_info.append(failed_entry)
                    # print(hp1)
                elif clpdata.startswith('NM1*') and '*82*' in clpdata:
                    cp4=clpdata.split('*')
                    provider_lastname,provider_firstname,provider_middlename,provider_prefix,provider_suffix,provider_id_code_qualifier,provider_id_code='notfound','notfound','notfound','notfound','notfound','notfound','notfound'
                    try:
                        provider_lastname=validation('provider_lastname',split_tilde(cp4[3]))
                        provider_firstname=validation('provider_firstname',split_tilde(cp4[4]))
                        provider_middlename=validation('provider_middlename',split_tilde(cp4[5]))
                        provider_prefix=validation('provider_prefix',split_tilde(cp4[6]))
                        provider_suffix=validation('provider_prefix',split_tilde(cp4[7]))
                        provider_id_code_qualifier=validation('provider_id_code_qualifier',split_tilde(cp4[8]))
                        provider_id_code=validation('provider_id_code',split_tilde(cp4[9]))
                    except:
                        pass

                    validation_results = {}

                    if provider_lastname != 'notfound':
                        validation_results['provider_lastname'] = (provider_lastname, split_tilde(cp4[3]), 3)
                    if provider_firstname != 'notfound':
                        validation_results['provider_firstname'] = (provider_firstname, split_tilde(cp4[4]), 4)
                    if provider_middlename != 'notfound':
                        validation_results['provider_middlename'] = (provider_middlename, split_tilde(cp4[5]), 5)
                    if provider_prefix != 'notfound':
                        validation_results['provider_prefix'] = (provider_prefix, split_tilde(cp4[6]), 6)
                    if provider_suffix != 'notfound':
                        validation_results['provider_suffix'] = (provider_suffix, split_tilde(cp4[7]), 7)
                    if provider_id_code_qualifier != 'notfound':
                        validation_results['provider_id_code_qualifier'] = (provider_id_code_qualifier, split_tilde(cp4[8]), 8)
                    if provider_id_code != 'notfound':
                        validation_results['provider_id_code'] = (provider_id_code, split_tilde(cp4[9]), 9)

                    failed_entry = [
                        {'lines': clpline, 'value': value, 'position': position, 'field': 'CLPData', 'title': var_name}
                        for var_name, (is_valid, value, position) in validation_results.items()
                        if not is_valid
                    ]

                    if failed_entry:
                        failed_info.append(failed_entry)

                clpline+=1
        svc_small=[]
        svc_lar=[]
        for index4,svc in enumerate(svc_chunks):
            svc_md=[]
            for index5,svc1 in enumerate(svc):
                if svc1.startswith('SVC*'):
                    svc_fd=True
                elif index5!=len(svc)-1:
                    if svc[index5+1].startswith('SVC*'):
                        svc_small.append(svc1)
                        svc_md.append(svc_small)
                        svc_fd=False
                        svc_small=[]
                elif index5==len(svc)-1:
                    svc_small.append(svc1)
                    svc_md.append(svc_small)
                    svc_fd=False
                    svc_small=[]
                if svc_fd:
                    svc_small.append(svc1)
                if not svc_fd and len(svc_small)>0:
                    svc_md.append(svc_small)
            svc_lar.append(svc_md)
        svc_line=header_count+len(clp_full[0])
        for index6,svcdata in enumerate(svc_lar):
            if index6>0:
                svc_line=svc_line+len(clp_full[index6])
            for index7,svcdata1 in enumerate(svcdata):
                for index8,svcdata2 in enumerate(svcdata1):
                    if svcdata2.startswith('SVC*'):
                        sd1=svcdata2.split('*')
                        procedure_code,product_service_id,billed,paid_amount='notfound','notfound','notfound','notfound'
                        try:
                            procedurecode=sd1[1]
                            if ':' in procedurecode:
                                proc_len=len(procedurecode.split(':'))
                                if proc_len==1:
                                    procedure_code=validation('procedure_code',split_tilde(procedurecode.split(':')[0]))
                                    product_service_id=validation('product_service_id',split_tilde(procedurecode.split(':')[1]))
                            else:
                                pass
                            
                            billed=validation('service_charge_amt',split_tilde(sd1[2]))
                            paid_amount=validation('paid_amount',split_tilde(sd1[3]))
                        except:
                            pass
                        validation_results = {}

                        if procedure_code != 'notfound':
                            validation_results['procedure_code'] = (procedure_code, split_tilde(sd1[1].split(':')[0]), 1)
                        if product_service_id != 'notfound':
                            validation_results['product_service_id'] = (product_service_id, split_tilde(sd1[1].split(':')[1]), 1)
                        if billed != 'notfound':
                            validation_results['service_charge_amt'] = (billed, split_tilde(sd1[2]), 2)
                        if paid_amount != 'notfound':
                            validation_results['paid_amount'] = (paid_amount, split_tilde(sd1[3]), 3)

                        failed_entry = [
                            {'lines': svc_line, 'value': value, 'position': position, 'field': 'ServiceLine', 'title': var_name}
                            for var_name, (is_valid, value, position) in validation_results.items()
                            if not is_valid
                        ]

                        if failed_entry:
                            failed_info.append(failed_entry)

                        # validation_results = {
                        # 'procedure_code': (procedure_code, split_tilde(procedurecode),1),
                        # 'billed': (billed, split_tilde(sd1[2]),2),
                        # 'service_pay_amount': (paid_amount, split_tilde(sd1[3]),3)
                        # }
                        # failed_entry=[
                        #     {'lines': svc_line, 'value': value,'position':position,'field':'SVC','title':var_name}
                        #     for var_name, (is_valid, value,position) in validation_results.items()
                        #     if not is_valid
                        # ]
                        # if failed_entry:
                        #     failed_info.append(failed_entry)
                    
                    elif svcdata2.startswith('DTM*'):
                        sd3=svcdata2.split('*')
                        service_date='notfound'
                        service_date_qualifier='notfound'
                        try:
                            service_date_qualifier=validation('service_date_qualifier',split_tilde(sd3[1]))
                            service_date=validation('pay_date',split_tilde(sd3[2]))
                        except:
                            pass
                        validation_results = {}

                        if service_date != 'notfound':
                            validation_results['service_date'] = (service_date, split_tilde(sd3[2]), 2)
                        if service_date_qualifier != 'notfound':
                            validation_results['service_date_qualifier'] = (service_date_qualifier, split_tilde(sd3[1]), 1)
                        

                        failed_entry = [
                            {'lines': svc_line, 'value': value, 'position': position, 'field': 'ServiceLine', 'title': var_name}
                            for var_name, (is_valid, value, position) in validation_results.items()
                            if not is_valid
                        ]

                        if failed_entry:
                            failed_info.append(failed_entry)

                    elif svcdata2.startswith('CAS*'):
                        sd2=svcdata2.split('*')
                        claim_grp_code,adjustment_reason_code,adjustment_amt='notfound','notfound','notfound'
                        try:
                            claim_grp_code=validation('claim_grp_code',split_tilde(sd2[1]))
                            adjustment_reason_code=validation('adjustment_reason_code',split_tilde(sd2[2]))
                            adjustment_amt=validation('adjustment_amt',split_tilde(sd2[3]))
                        except:
                            pass

                        validation_results = {}

                        if claim_grp_code != 'notfound':
                            validation_results['claim_grp_code'] = (claim_grp_code, split_tilde(sd2[1]), 1)
                        if adjustment_reason_code != 'notfound':
                            validation_results['adjustment_reason_code'] = (adjustment_reason_code, split_tilde(sd2[2]), 2)
                        if adjustment_amt != 'notfound':
                            validation_results['adjustment_amt'] = (adjustment_amt, split_tilde(sd2[3]), 3)

                        failed_entry = [
                            {'lines': svc_line, 'value': value, 'position': position, 'field': 'ServiceLine', 'title': var_name}
                            for var_name, (is_valid, value, position) in validation_results.items()
                            if not is_valid
                        ]

                        if failed_entry:
                            failed_info.append(failed_entry)
                    # elif svcdata2.startswith('AMT*'):
                

                    svc_line+=1
        summary_lines=[]
        summary_fd=False
        for index9,summ in enumerate(full_lines):
            if index9==totalcount2:
                summary_fd=True
            if summary_fd:
                summary_lines.append(full_lines[index9])
        summary_lines_count=totalcount2
        for index10,summary in enumerate(summary_lines):
            if summary.startswith('SE*'):
                sud1=summary.split('*')
                included_sgmt_no,trans_set_control_number='notfound','notfound'
                try:
                    included_sgmt_no=validation('included_sgmt_no',split_tilde(sud1[1]))
                    trans_set_control_number=validation('trans_set_control_number',split_tilde(sud1[2]))
                except:
                    pass
                validation_results = {}

                if included_sgmt_no != 'notfound':
                    validation_results['included_sgmt_no'] = (included_sgmt_no, split_tilde(sud1[1]), 1)
                if trans_set_control_number != 'notfound':
                    validation_results['trans_set_control_number'] = (trans_set_control_number, split_tilde(sud1[2]), 2)
                failed_entry = [
                    {'lines': summary_lines_count, 'value': value, 'position': position, 'field': 'SummaryLine', 'title': var_name}
                    for var_name, (is_valid, value, position) in validation_results.items()
                    if not is_valid
                ]

                if failed_entry:
                    failed_info.append(failed_entry)
            elif summary.startswith('GE*'):
                sud2=summary.split('*')
                number_of_transaction_set,group_control_no='notfound','notfound'
                try:
                    number_of_transaction_set=validation('number_of_transaction_set',split_tilde(sud2[1]))
                    group_control_no=validation('group_control_no',split_tilde(sud2[2]))
                except:
                    pass
                validation_results = {}
                if number_of_transaction_set != 'notfound':
                    validation_results['number_of_transaction_set'] = (number_of_transaction_set, split_tilde(sud2[1]), 1)
                if group_control_no != 'notfound':
                    validation_results['group_control_no'] = (group_control_no, split_tilde(sud2[2]), 2)
                failed_entry = [
                    {'lines': summary_lines_count, 'value': value, 'position': position, 'field': 'SummaryLine', 'title': var_name}
                    for var_name, (is_valid, value, position) in validation_results.items()
                    if not is_valid
                ]

                if failed_entry:
                    failed_info.append(failed_entry)
            elif summary.startswith('IEA*'):
                no_of_included_segments,interchange_control_no='notfound','notfound'
                sud3=summary.split('*')
                try:
                    no_of_included_segments=validation('no_of_included_segments',split_tilde(sud3[1]))
                    interchange_control_no=validation('interchange_control_no',split_tilde(sud3[2]))
                except:
                    pass
                validation_results={}
                if no_of_included_segments != 'notfound':
                    validation_results['no_of_included_segments'] = (no_of_included_segments, split_tilde(sud3[1]), 1)
                if interchange_control_no != 'notfound':
                    validation_results['interchange_control_no'] = (interchange_control_no, split_tilde(sud3[2]), 2)
                failed_entry = [
                    {'lines': summary_lines_count, 'value': value, 'position': position, 'field': 'SummaryLine', 'title': var_name}
                    for var_name, (is_valid, value, position) in validation_results.items()
                    if not is_valid
                ]
                if failed_entry:
                    failed_info.append(failed_entry)
            summary_lines_count+=1

        final_failed_info.append(failed_info)

        # pass

    # print(failed_info)
    # os.remove(new_path)
    process_files(final_failed_info,excel_path)
    return True


class CustomButton(QPushButton):
    def __init__(self, text, icon_path=None, tooltip=None):
        super().__init__(text)
        self.setMinimumHeight(40)
        self.setFont(QFont("Roboto", 10))
        self.setCursor(Qt.PointingHandCursor)
        self.setStyleSheet("""
            QPushButton {
                background-color: #4A90E2;
                color: white;
                border-radius: 6px;
                padding: 10px 15px;
                border: none;
            }
            QPushButton:hover {
                background-color: #357ABD;
            }
            QPushButton:pressed {
                background-color: #2A6395;
            }
            QPushButton:disabled {
                background-color: #A0A0A0;
                color: #D0D0D0;
            }
        """)
        if icon_path and os.path.exists(icon_path):
            self.setIcon(QIcon(icon_path))
            self.setIconSize(QSize(20, 20))
        if tooltip:
            self.setToolTip(tooltip)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("835 Validation Tool")
        self.setGeometry(100, 100, 1000, 700)
        self.setStyleSheet("""
            QMainWindow {
                background-color: #F7F9FC;
            }
            QWidget {
                font-family: 'Roboto', sans-serif;
                color: #2D3748;
            }
            QTextEdit {
                background-color: white;
                border: none;
                border-radius: 8px;
                font-family: 'Courier New', monospace;
                font-size: 9pt;
                padding: 10px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            }
            QLineEdit {
                border: 1px solid #E2E8F0;
                border-radius: 6px;
                padding: 8px 12px;
                background-color: white;
                font-size: 10pt;
            }
            QLineEdit:focus {
                border: 2px solid #4A90E2;
                box-shadow: 0 0 5px rgba(74,144,226,0.3);
            }
            QStatusBar {
                background-color: #EDF2F7;
                color: #4A5568;
                border-top: 1px solid #E2E8F0;
                padding: 5px;
            }
            QFrame#sidebar {
                background-color: #2D3748;
            }
            QLabel#app_title {
                color: white;
                font-size: 20pt;
                font-weight: 500;
            }
            QLabel#sidebar_label {
                color: #A0AEC0;
                font-size: 11pt;
                font-weight: 500;
                text-transform: uppercase;
                letter-spacing: 1px;
            }
        """)

        logo_path = r"D:\py files\Call App\droidal-logo.png"
        if os.path.exists(logo_path):
            self.setWindowIcon(QIcon(logo_path))

        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        self.sidebar = QFrame()
        self.sidebar.setObjectName("sidebar")
        self.sidebar.setMinimumWidth(280)
        self.sidebar.setMaximumWidth(280)
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(20, 30, 20, 30)
        sidebar_layout.setSpacing(20)

        title_layout = QHBoxLayout()
        # if os.path.exists(logo_path):
        #     logo_label = QLabel()
        #     pixmap = QPixmap(logo_path).scaled(40, 40, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        #     logo_label.setPixmap(pixmap)
        #     title_layout.addWidget(logo_label)
        
        app_title = QLabel("835 Validator")
        app_title.setObjectName("app_title")
        title_layout.addWidget(app_title)
        title_layout.addStretch()
        sidebar_layout.addLayout(title_layout)
        
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setStyleSheet("background-color: #4A5568; margin: 10px 0;")
        separator.setMaximumHeight(1)
        sidebar_layout.addWidget(separator)
        
        mode_label = QLabel("Validation Mode")
        mode_label.setObjectName("sidebar_label")
        sidebar_layout.addWidget(mode_label)

        self.single_btn = CustomButton("Single File Validation", tooltip="Validate a single 835 file")
        self.multi_btn = CustomButton("Multiple File Validation", tooltip="Validate multiple 835 files")
        
        self.single_btn.setStyleSheet(self.single_btn.styleSheet() + """
            QPushButton {
                background-color: #4A5568;
                padding: 12px;
                text-align: left;
                margin-bottom: 10px;
            }
            QPushButton:hover {
                background-color: #718096;
            }
        """)
        
        self.multi_btn.setStyleSheet(self.multi_btn.styleSheet() + """
            QPushButton {
                background-color: #4A5568;
                padding: 12px;
                text-align: left;
            }
            QPushButton:hover {
                background-color: #718096;
            }
        """)
        
        sidebar_layout.addWidget(self.single_btn)
        sidebar_layout.addWidget(self.multi_btn)
        
        actions_label = QLabel("Actions")
        actions_label.setObjectName("sidebar_label")
        sidebar_layout.addWidget(actions_label)
        
        self.refresh_btn = CustomButton("Refresh File", tooltip="Refresh current file")
        self.refresh_btn.setStyleSheet(self.refresh_btn.styleSheet() + """
            QPushButton {
                background-color: #4A5568;
                padding: 12px;
                text-align: left;
            }
            QPushButton:hover {
                background-color: #718096;
            }
        """)
        self.refresh_btn.setVisible(False)
        
        sidebar_layout.addWidget(self.refresh_btn)
        sidebar_layout.addStretch()
        
        help_btn = CustomButton("Help")
        help_btn.setStyleSheet(help_btn.styleSheet() + """
            QPushButton {
                background-color: #718096;
            }
            QPushButton:hover {
                background-color: #A0AEC0;
            }
        """)
        sidebar_layout.addWidget(help_btn)
        
        main_layout.addWidget(self.sidebar)
        
        content_widget = QWidget()
        content_widget.setStyleSheet("""
            QWidget {
                background-color: #F7F9FC;
                border-radius: 8px;
            }
        """)
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(25, 25, 25, 25)
        content_layout.setSpacing(20)
        
        header_layout = QHBoxLayout()
        self.content_title = QLabel("Welcome")
        self.content_title.setStyleSheet("""
            font-size: 18pt;
            font-weight: 500;
            color: #2D3748;
        """)
        
        self.toggle_size_btn = QPushButton("Maximize")
        self.toggle_size_btn.setMinimumHeight(40)
        self.toggle_size_btn.setMaximumWidth(120)
        self.toggle_size_btn.setCursor(Qt.PointingHandCursor)
        self.toggle_size_btn.setStyleSheet("""
            QPushButton {
                background-color: #EDF2F7;
                color: #4A5568;
                border: 1px solid #E2E8F0;
                border-radius: 4px;
                padding: 4px 10px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #E2E8F0;
            }
        """)
        self.toggle_size_btn.clicked.connect(self.toggle_content_size)
        self.toggle_size_btn.setVisible(False)
        
        header_layout.addWidget(self.content_title)
        header_layout.addStretch()
        header_layout.addWidget(self.toggle_size_btn)
        content_layout.addLayout(header_layout)
        
        path_layout = QHBoxLayout()
        self.path_input = QLineEdit()
        self.path_input.setVisible(False)
        self.path_input.setMinimumHeight(40)
        self.path_input.setPlaceholderText("Enter file path...")
        self.path_input.returnPressed.connect(self.process_input)
        
        self.browse_btn = QPushButton("Browse")
        self.browse_btn.setMinimumHeight(40)
        self.browse_btn.setMaximumWidth(120)
        self.browse_btn.setCursor(Qt.PointingHandCursor)
        self.browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #EDF2F7;
                color: #4A5568;
                border: 1px solid #E2E8F0;
                border-radius: 6px;
                padding: 8px 15px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #E2E8F0;
            }
        """)
        self.browse_btn.setVisible(False)
        self.browse_btn.clicked.connect(self.browse_file)
        
        path_layout.addWidget(self.path_input)
        path_layout.addSpacing(10)
        path_layout.addWidget(self.browse_btn)
        content_layout.addLayout(path_layout)
        
        self.status_layout = QHBoxLayout()
        self.status_indicator = QLabel()
        self.status_indicator.setVisible(False)
        self.status_indicator.setStyleSheet("""
            padding: 6px 12px;
            border-radius: 15px;
            background-color: #EDF2F7;
            font-weight: 500;
        """)
        self.status_layout.addWidget(self.status_indicator)
        self.status_layout.addStretch()
        content_layout.addLayout(self.status_layout)
        
        display_widget = QWidget()
        display_widget.setStyleSheet("""
            background-color: white;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        """)
        self.display_layout = QHBoxLayout(display_widget)
        self.display_layout.setContentsMargins(15, 15, 15, 15)
        
        self.text_display = QTextEdit()
        self.text_display.setReadOnly(True)
        self.text_display.setLineWrapMode(QTextEdit.NoWrap)
        
        self.info_panel = QTextEdit()
        self.info_panel.setReadOnly(True)
        self.info_panel.setMaximumWidth(350)
        self.info_panel.setStyleSheet("""
            QTextEdit {
                background-color: #F7F9FC;
                border-left: 1px solid #E2E8F0;
                border-radius: 0;
                box-shadow: none;
                padding: 15px;
            }
        """)
        self.info_panel.setHtml("""
            <h3 style="color: #2D3748; margin-bottom: 15px;">Validation Details</h3>
            <p style="color: #718096;">No file loaded yet.</p>
            <p style="color: #718096;">Select a validation mode to begin.</p>
        """)
        
        self.display_layout.addWidget(self.text_display, 2)
        self.display_layout.addWidget(self.info_panel, 1)
        
        content_layout.addWidget(display_widget, 1)
        main_layout.addWidget(content_widget)
        
        self.statusBar = QStatusBar()
        self.statusBar.showMessage("Ready")
        self.setStatusBar(self.statusBar)
        
        self.single_btn.clicked.connect(self.single_file_mode)
        self.multi_btn.clicked.connect(self.multi_file_mode)
        self.refresh_btn.clicked.connect(self.refresh_single_file)
        help_btn.clicked.connect(self.show_help)
        
        self.current_mode = None
        self.excel_path = None
        self.current_file = None
        self.is_content_maximized = False
        self.file_watcher = QFileSystemWatcher()
        self.file_watcher.fileChanged.connect(self.on_file_changed)
        
        self.text_display.setHtml("""
            <div style="margin: 50px; text-align: center;">
                <h1 style="color: #2D3748; font-size: 24px; margin-bottom: 20px;">Welcome to 835 Validation Tool</h1>
                <p style="color: #718096; font-size: 14pt; line-height: 1.5;">
                    Select a validation mode from the sidebar to begin processing your 835 files.
                </p>
                <div style="margin-top: 30px; text-align: left; max-width: 500px; margin-left: auto; margin-right: auto;">
                    <p><b style="color: #4A90E2;">Single File Mode:</b> Validate and highlight errors in a single file</p>
                    <p><b style="color: #4A90E2;">Multiple File Mode:</b> Process multiple files with Excel export</p>
                </div>
            </div>
        """)

    def toggle_content_size(self):
        if not self.is_content_maximized:
            self.sidebar.setVisible(False)
            self.info_panel.setVisible(False)
            self.path_input.setVisible(False)
            self.browse_btn.setVisible(False)
            self.refresh_btn.setVisible(False)
            self.status_indicator.setVisible(False)
            self.display_layout.setStretch(0, 1)
            self.toggle_size_btn.setText("Restore")
            self.statusBar.showMessage("Content maximized")
            self.is_content_maximized = True
        else:
            self.sidebar.setVisible(True)
            self.info_panel.setVisible(True)
            self.path_input.setVisible(self.current_mode is not None)
            self.browse_btn.setVisible(self.current_mode is not None)
            self.refresh_btn.setVisible(self.current_file is not None)
            self.status_indicator.setVisible(self.current_file is not None)
            self.display_layout.setStretch(0, 2)
            self.display_layout.setStretch(1, 1)
            self.toggle_size_btn.setText("Maximize")
            self.statusBar.showMessage("Content restored")
            self.is_content_maximized = False

    def single_file_mode(self):
        self.current_mode = "single"
        self.content_title.setText("Single File Validation")
        self.path_input.setPlaceholderText("Enter file path or browse...")
        self.path_input.setVisible(True)
        self.browse_btn.setVisible(True)
        self.refresh_btn.setVisible(False)
        self.toggle_size_btn.setVisible(False)
        self.path_input.clear()
        self.text_display.clear()
        self.status_indicator.setVisible(False)
        self.info_panel.setHtml("""
            <h3 style="color: #2D3748; margin-bottom: 15px;">Single File Validation</h3>
            <p style="color: #718096;">Please load a 835 file to validate.</p>
            <p style="color: #718096;">Validation errors will be highlighted in red.</p>
        """)
        self.statusBar.showMessage("Single file mode activated")

    def multi_file_mode(self):
        self.current_mode = "multiple"
        self.content_title.setText("Multiple File Validation")
        self.path_input.setVisible(False)  # Hide path input since we'll use dialogs directly
        self.browse_btn.setVisible(False)
        self.refresh_btn.setVisible(False)
        self.toggle_size_btn.setVisible(False)
        self.text_display.clear()
        self.status_indicator.setVisible(False)
        self.info_panel.setHtml("""
            <h3 style="color: #2D3748; margin-bottom: 15px;">Multiple File Validation</h3>
            <p style="color: #718096;">You will be prompted to select:</p>
            <p style="color: #718096;">1. Folder containing 835 files</p>
            <p style="color: #718096;">2. Excel output file location</p>
        """)
        self.statusBar.showMessage("Multiple file mode activated")
        self.process_input()  # Trigger the process immediately

    def browse_file(self):
        if self.current_mode == "single":
            filepath, _ = QFileDialog.getOpenFileName(
                self, "Select 835 File", "", "Text Files (*.txt);;All Files (*)"
            )
            if filepath:
                self.path_input.setText(filepath)
                self.process_input()

    def process_input(self):
        if self.current_mode == "single":
            path = self.path_input.text().strip()
            if not path:
                return
            self.process_single_file(path)
        elif self.current_mode == "multiple":
            # Step 1: Select folder containing 835 files
            folder_path = QFileDialog.getExistingDirectory(
                self, "Select Folder Containing 835 Files"
            )
            if not folder_path:
                return
            
            # Step 2: Select Excel output file
            excel_path = QFileDialog.getSaveFileName(
                self, "Select Excel Output File", "", "Excel Files (*.xlsx)"
            )[0]
            if not excel_path:
                return
                
            self.process_multiple_files(folder_path, excel_path)

    def process_single_file(self, filepath):
        try:
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"File not found: {filepath}")
            
            self.statusBar.showMessage(f"Processing file: {os.path.basename(filepath)}...")
            self.status_indicator.setText("⏳ Processing...")
            self.status_indicator.setStyleSheet("""
                padding: 6px 12px;
                border-radius: 15px;
                background-color: #EBF8FF;
                color: #4A90E2;
                font-weight: 500;
            """)
            self.status_indicator.setVisible(True)
            QApplication.processEvents()
            
            with open(filepath, 'r', newline='') as file:
                content = file.read()
            
            lines = content.split('~')
            lines = [line.strip() for line in lines if line.strip()]
            content = '\n'.join(lines)
            
            self.text_display.setText(content)
            self.current_file = filepath
            
            if self.current_file in self.file_watcher.files():
                self.file_watcher.removePath(self.current_file)
            self.file_watcher.addPath(self.current_file)
            
            results = validation_singlefile(self.current_file)
            self.update_highlights(results)
            self.refresh_btn.setVisible(True)
            self.toggle_size_btn.setVisible(True)
            
            error_count = sum(len(group) for group in results)
            if error_count > 0:
                self.status_indicator.setText(f"⚠️ {error_count} Issues Found")
                self.status_indicator.setStyleSheet("""
                    padding: 6px 12px;
                    border-radius: 15px;
                    background-color: #FFF5F5;
                    color: #E53E3E;
                    font-weight: 500;
                """)
            else:
                self.status_indicator.setText("✅ Validated Successfully")
                self.status_indicator.setStyleSheet("""
                    padding: 6px 12px;
                    border-radius: 15px;
                    background-color: #F0FFF4;
                    color: #38A169;
                    font-weight: 500;
                """)
            
            self.update_info_panel(results)
            self.statusBar.showMessage(f"File loaded: {os.path.basename(filepath)}")
            
        except Exception as e:
            self.status_indicator.setText("❌ Error")
            self.status_indicator.setStyleSheet("""
                padding: 6px 12px;
                border-radius: 15px;
                background-color: #FFF5F5;
                color: #E53E3E;
                font-weight: 500;
            """)
            QMessageBox.critical(self, "Error", f"Error processing file: {str(e)}")
            self.statusBar.showMessage("Error processing file")

    def update_highlights(self, results=None):
        if self.current_file:
            if results is None:
                results = validation_singlefile(self.current_file)
            
            self.text_display.clear()
            with open(self.current_file, 'r', newline='') as file:
                content = file.read()
            lines = content.split('~')
            lines = [line.strip() for line in lines if line.strip()]
            content = '\n'.join(lines)
            self.text_display.setText(content)
            
            for result_group in results:
                for item in result_group:
                    self.highlight_text(item['lines'], item['position'], item['field'], item['title'], item['value'])

    def highlight_text(self, line_num, position, field, title, value):
        cursor = self.text_display.textCursor()
        cursor.movePosition(QTextCursor.Start)
        for i in range(line_num):
            cursor.movePosition(QTextCursor.Down)
        
        cursor.select(QTextCursor.LineUnderCursor)
        line_text = cursor.selectedText()
        segments = line_text.split('*')
        
        if position < len(segments):
            char_pos = 0
            for i in range(position):
                char_pos += len(segments[i]) + 1
            
            format = QTextCharFormat()
            format.setBackground(QColor(255, 245, 245))
            format.setForeground(QColor(229, 62, 62))
            format.setToolTip(f"Field: {field} > {title}\nExpected value: {value}")
            
            cursor.movePosition(QTextCursor.StartOfLine)
            cursor.movePosition(QTextCursor.Right, QTextCursor.MoveAnchor, char_pos)
            
            if segments[position] == '':
                cursor.movePosition(QTextCursor.Right, QTextCursor.KeepAnchor, 1)
            else:
                cursor.movePosition(QTextCursor.Right, QTextCursor.KeepAnchor, len(segments[position]))
            
            cursor.mergeCharFormat(format)

    def update_info_panel(self, results):
        error_count = sum(len(group) for group in results)
        
        html_content = f"""
        <h3 style="color: #2D3748; margin-bottom: 15px;">Validation Results</h3>
        <p><b>File:</b> {os.path.basename(self.current_file)}</p>
        <p><b>Status:</b> <span style="color:{'#E53E3E' if error_count > 0 else '#38A169'}">
        {'❌ Failed' if error_count > 0 else '✅ Passed'}</span></p>
        <p><b>Issues:</b> {error_count}</p>
        <hr style="border: 0; border-top: 1px solid #E2E8F0; margin: 15px 0;">
        """
        
        if error_count > 0:
            html_content += "<h4 style='margin-bottom: 10px;'>Details:</h4><ul style='margin-left: -20px;'>"
            for group_index, result_group in enumerate(results):
                for item in result_group:
                    field_info = f"{item['field']} > {item['title']}"
                    value_info = f"Value: '{item['value']}'" if item['value'] else "Missing value"
                    html_content += f"""
                    <li style="margin-bottom: 12px; color: #718096;">
                        <div><b style="color: #2D3748;">{field_info}</b></div>
                        <div style="color: #E53E3E;">{value_info}</div>
                    </li>
                    """
            html_content += "</ul>"
        else:
            html_content += "<p style='color: #718096;'>No validation issues found.</p>"
        
        self.info_panel.setHtml(html_content)

    def on_file_changed(self, path):
        if path == self.current_file:
            self.statusBar.showMessage(f"File changed: {os.path.basename(path)}")
            self.update_highlights()

    def refresh_single_file(self):
        if self.current_file:
            self.statusBar.showMessage(f"Refreshing file: {os.path.basename(self.current_file)}")
            self.update_highlights()
            self.toggle_size_btn.setVisible(True)
        else:
            QMessageBox.warning(self, "Warning", "No file loaded to refresh.")

    def process_multiple_files(self, folder_path, excel_path):
        try:
            if not os.path.exists(folder_path):
                raise FileNotFoundError(f"Folder not found: {folder_path}")
                
            self.statusBar.showMessage("Processing multiple files...")
            self.status_indicator.setText("⏳ Processing...")
            self.status_indicator.setStyleSheet("""
                padding: 6px 12px;
                border-radius: 15px;
                background-color: #EBF8FF;
                color: #4A90E2;
                font-weight: 500;
            """)
            self.status_indicator.setVisible(True)
            QApplication.processEvents()
            
            result = validation_multiplefile(folder_path, excel_path)
            
            if result:
                self.status_indicator.setText("✅ Processing Complete")
                self.status_indicator.setStyleSheet("""
                    padding: 6px 12px;
                    border-radius: 15px;
                    background-color: #F0FFF4;
                    color: #38A169;
                    font-weight: 500;
                """)
                QMessageBox.information(self, "Success", f"Data saved to:\n{excel_path}")
                
                self.info_panel.setHtml(f"""
                    <h3 style="color: #2D3748; margin-bottom: 15px;">Multiple File Validation</h3>
                    <p><b>Status:</b> <span style="color:#38A169">✅ Completed</span></p>
                    <p><b>Folder:</b> {folder_path}</p>
                    <p><b>Output:</b> {os.path.basename(excel_path)}</p>
                    <p style="color: #718096;">All files processed successfully.</p>
                """)
            else:
                self.status_indicator.setText("❌ Processing Failed")
                self.status_indicator.setStyleSheet("""
                    padding: 6px 12px;
                    border-radius: 15px;
                    background-color: #FFF5F5;
                    color: #E53E3E;
                    font-weight: 500;
                """)
                
            self.path_input.setVisible(False)
            self.browse_btn.setVisible(False)
            self.toggle_size_btn.setVisible(False)
            self.excel_path = None
            self.statusBar.showMessage("Processing complete")
            
        except Exception as e:
            self.status_indicator.setText("❌ Error")
            self.status_indicator.setStyleSheet("""
                padding: 6px 12px;
                border-radius: 15px;
                background-color: #FFF5F5;
                color: #E53E3E;
                font-weight: 500;
            """)
            QMessageBox.critical(self, "Error", f"Error processing files: {str(e)}")
            self.statusBar.showMessage("Error processing files")

    def show_help(self):
        help_text = """
        <h2 style="color: #2D3748; margin-bottom: 20px;">835 Validation Tool Help</h2>
        
        <h3 style="color: #4A5568;">Single File Validation</h3>
        <p style="color: #718096; line-height: 1.5;">Validate a single 835 file with visual error highlighting.</p>
        <ol style="color: #718096; margin-left: 20px;">
            <li>Select "Single File Validation"</li>
            <li>Enter path or browse for file</li>
            <li>View highlighted errors</li>
            <li>Use "Refresh File" to reload</li>
        </ol>
        
        <h3 style="color: #4A5568;">Multiple File Validation</h3>
        <p style="color: #718096; line-height: 1.5;">Process multiple 835 files with Excel export.</p>
        <ol style="color: #718096; margin-left: 20px;">
            <li>Select "Multiple File Validation"</li>
            <li>Select folder with 835 files</li>
            <li>Choose output Excel file location</li>
            <li>Wait for processing completion</li>
        </ol>
        
        <h3 style="color: #4A5568;">Validation Results</h3>
        <p style="color: #718096; line-height: 1.5;">Errors are highlighted in red with details in the right panel.</p>
        """
        
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Help")
        msg_box.setTextFormat(Qt.RichText)
        msg_box.setText(help_text)
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #FFFFFF;
            }
            QLabel {
                color: #2D3748;
            }
        """)
        msg_box.exec_()

if __name__ == '__main__':
    QCoreApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QCoreApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    
    app = QApplication(sys.argv)
    app.setFont(QFont("Roboto", 10))
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
